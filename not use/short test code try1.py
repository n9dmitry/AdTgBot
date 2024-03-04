from aiogram import Bot, Dispatcher, types, executor
from aiogram.types import InputMediaPhoto
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton
from aiogram.utils.markdown import hlink
import random
import datetime
import uuid
import asyncio
import openpyxl
from fuzzywuzzy import fuzz
from config import *
from states import *
import json
from enumlist import *

# –ó–∞–≥—Ä—É–∑–∫–∞ JSON –≤ –Ω–∞—á–∞–ª–µ —Å–∫—Ä–∏–ø—Ç–∞
with open('../dicts.json', 'r', encoding='utf-8') as file:
    dicts = json.load(file)

dict_start_brands = dicts.get("dict_start_brands", {})
dict_car_brands_and_models = dicts.get("dict_car_brands_and_models", {})
dict_car_body_types = dicts.get("dict_car_body_types", {})
dict_car_engine_types = dicts.get("dict_car_engine_types", {})
dict_car_transmission_types = dicts.get("dict_car_transmission_types", {})
dict_car_colors = dicts.get("dict_car_colors", {})
dict_car_document_statuses = dicts.get("dict_car_document_statuses", {})
dict_car_owners = dicts.get("dict_car_owners", {})
dict_car_customs_cleared = dicts.get("dict_car_customs_cleared", {})
dict_currency = dicts.get("dict_currency", {})
dict_car_conditions = dicts.get("dict_car_conditions", {})
dict_car_mileages = dicts.get("dict_car_mileages", {})
dict_edit_buttons = dicts.get("dict_edit_buttons", {})


# –ö–æ–Ω–µ—Ü –∏–º–ø–æ—Ä—Ç–∞ json —Å–ª–æ–≤–∞—Ä–µ–π


# –°–æ–∑–¥–∞–Ω–∏–µ –∫–ª–∞–≤–∏–∞—Ç—É—Ä—ã
def create_keyboard(button_texts, resize_keyboard=True):
    keyboard = ReplyKeyboardMarkup(
        resize_keyboard=resize_keyboard, row_width=2)
    buttons = [KeyboardButton(text=text) for text in button_texts]
    keyboard.add(*buttons)
    return keyboard


async def recognize_car_model(event, brand_name):
    models = []
    similar_brands = []

    if brand_name.lower() in ['–∂–∏–≥—É–ª–∏']:
        brand_name = 'Lada (–í–ê–ó)'

    with open('../cars.json', encoding='utf-8') as file:
        data = json.load(file)

    found_brand = False
    for item in data:
        # –°—Ä–∞–≤–Ω–µ–Ω–∏–µ –∏–º–µ–Ω–∏ –±—Ä–µ–Ω–¥–∞ —Å —É—á–µ—Ç–æ–º —Ä–∞—Å—Å—Ç–æ—è–Ω–∏—è –õ–µ–≤–µ–Ω—à—Ç–µ–π–Ω–∞
        if 'name' in item and fuzz.token_sort_ratio(brand_name.lower(), item['name'].lower()) >= 90:
            if 'models' in item:
                models = item['models']
            found_brand = True
            break
        # –°—Ä–∞–≤–Ω–µ–Ω–∏–µ –∫–∏—Ä–∏–ª–ª–∏—á–µ—Å–∫–æ–≥–æ –∏–º–µ–Ω–∏ –±—Ä–µ–Ω–¥–∞
        elif 'cyrillic-name' in item and fuzz.token_sort_ratio(brand_name.lower(), item['cyrillic-name'].lower()) >= 90:
            if 'models' in item:
                models = item['models']
            found_brand = True
            break

    if not found_brand and len(brand_name) >= 3:
        for inner_item in data:
            # –ü–æ–∏—Å–∫ –ø–æ—Ö–æ–∂–∏—Ö –±—Ä–µ–Ω–¥–æ–≤ —Å —É—á–µ—Ç–æ–º —Ä–∞—Å—Å—Ç–æ—è–Ω–∏—è –õ–µ–≤–µ–Ω—à—Ç–µ–π–Ω–∞
            if 'name' in inner_item and fuzz.token_sort_ratio(brand_name.lower(), inner_item['name'].lower()) >= 50 \
                and inner_item['name'] not in similar_brands:
                similar_brands.append(inner_item['name'])
            # –ü–æ–∏—Å–∫ –ø–æ—Ö–æ–∂–∏—Ö –∫–∏—Ä–∏–ª–ª–∏—á–µ—Å–∫–∏—Ö –±—Ä–µ–Ω–¥–æ–≤
            elif 'cyrillic-name' in inner_item and fuzz.token_sort_ratio(brand_name.lower(), inner_item['cyrillic-name'].lower()) >= 50 \
                and inner_item['name'] not in similar_brands:
                similar_brands.append(inner_item['name'])

        if similar_brands:
            response_message = "–ü–æ—Ö–æ–∂–∏–µ –±—Ä–µ–Ω–¥—ã:\n" + "\n".join(similar_brands)
            await event.answer(response_message)
    return models


class CarBotHandler:
    def __init__(self):
        self.lock = asyncio.Lock()

    # –ö–æ–º–∞–Ω–¥—ã

    async def restart(self, message, state):
        await state.finish()
        await message.answer("–ë–æ—Ç –ø–µ—Ä–µ–∑–∞–ø—É—â–µ–Ω.")
        await self.start(message, state)

    async def support(self, message, state):
        await state.finish()
        self.secret_number = str(random.randint(100, 999))

        await message.answer(f"–ù–∞—à–ª–∏ –±–∞–≥? –î–∞–≤–∞–π—Ç–µ –æ—Ç–ø—Ä–∞–≤–∏–º —Å–æ–æ–±—â–µ–Ω–∏–µ —Ä–∞–∑—Ä–∞–±–æ—Ç—á–∏–∫–∞–º! "
                             f"–ù–æ –ø–µ—Ä–µ–¥ —ç—Ç–∏–º –≤–≤–µ–¥–∏—Ç–µ –ø—Ä–æ–≤–µ—Ä–∫—É. –î–æ–∫–∞–∂–∏—Ç–µ —á—Ç–æ –≤—ã –Ω–µ —Ä–æ–±–æ—Ç. –ù–∞–ø–∏—à–∏—Ç–µ —á–∏—Å–ª–æ {self.secret_number}:")
        await state.set_state(User.STATE_SUPPORT_VALIDATION)

    async def support_validation(self, message, state):
        if message.text.isdigit() and message.text == self.secret_number:
            await message.reply(f"–ü—Ä–æ–≤–µ—Ä–∫–∞ –ø—Ä–æ–π–¥–µ–Ω–∞ —É—Å–ø–µ—à–Ω–æ!")
            await asyncio.sleep(1)
            await message.answer(f"–û–ø–∏—à–∏—Ç–µ —Ç–µ—Ö–Ω–∏—á–µ—Å–∫—É—é –ø—Ä–æ–±–ª–µ–º—É –≤ –¥–µ—Ç–∞–ª—è—Ö –¥–ª—è —Ä–∞–∑—Ä–∞–±–æ—Ç—á–∏–∫–æ–≤: ")
            await state.set_state(User.STATE_SUPPORT_MESSAGE)
        else:
            await message.answer(f"–ü–æ–ø—Ä–æ–±—É–π—Ç–µ –µ—â—ë —Ä–∞–∑!")
            await asyncio.sleep(1)
            await cmd_support(message, state)

    async def support_message(self, message: types.Message, state):
        # –ü–æ–ª—É—á–∞–µ–º —Ç–µ–∫—É—â—É—é –¥–∞—Ç—É –∏ –≤—Ä–µ–º—è
        current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        # –§–æ—Ä–º–∏—Ä—É–µ–º —Å—Ç—Ä–æ–∫—É –¥–ª—è –∑–∞–ø–∏—Å–∏ –≤ —Ñ–∞–π–ª
        message_to_write = f"""
        –î–∞—Ç–∞: {current_time}
        –ò–º—è: {message.from_user.full_name}
        Telegram @{message.from_user.username or message.from_user.id} 

        –°–æ–æ–±—â–µ–Ω–∏–µ: {message.text}
        ...
            """

        # –û—Ç–∫—Ä—ã–≤–∞–µ–º —Ñ–∞–π–ª –¥–ª—è –∑–∞–ø–∏—Å–∏ –∏ –∑–∞–ø–∏—Å—ã–≤–∞–µ–º —Å–æ–æ–±—â–µ–Ω–∏–µ
        with open("support.txt", "a") as file:
            file.write(message_to_write)
        keyboard = create_keyboard(['–ü–µ—Ä–µ–∑–∞–≥—Ä—É–∑–∏—Ç—å –±–æ—Ç–∞'])
        await message.reply("–°–ø–∞—Å–∏–±–æ –∑–∞ –≤–∞—à–µ —Å–æ–æ–±—â–µ–Ω–∏–µ! –ú—ã —Ä–∞—Å—Å–º–æ—Ç—Ä–∏–º –≤–∞—à—É –ø—Ä–æ–±–ª–µ–º—É!", reply_markup=keyboard)
        await state.set_state(User.STATE_SUPPORT_END)

    async def support_end(selfself, message, state):
        if message.text == '–ü–µ—Ä–µ–∑–∞–≥—Ä—É–∑–∏—Ç—å –±–æ—Ç–∞':
            await cmd_restart(message, state)
        await state.finish()

    async def start(self, message, state):
        image_hello_path = ImageDirectory.auto_say_hi
        with open(image_hello_path, "rb") as image_hello:
            self.m = await message.answer_photo(image_hello,
                                                caption=f"–ü—Ä–∏–≤–µ—Ç, {message.from_user.first_name}! –î–∞–≤–∞–π –ø—Ä–æ–¥–∞–¥–∏–º —Ç–≤–æ—ë –∞–≤—Ç–æ! –ù–∞—á–Ω—ë–º –∂–µ —Å–±–æ—Ä –¥–∞–Ω–Ω—ã—Ö!")
        await asyncio.sleep(0)

        keyboard = create_keyboard(dict_start_brands)

        image_path = ImageDirectory.auto_car_brand  # –ü—É—Ç—å –∫ –≤–∞—à–µ–º—É –∏–∑–æ–±—Ä–∞–∂–µ–Ω–∏—é
        with open(image_path, "rb") as image:
            await message.answer_photo(image, caption="–í—ã–±–µ—Ä–∏—Ç–µ –æ–¥–Ω—É –∏–∑ –∫–Ω–æ–ø–æ–∫ –Ω–∏–∂–µ –∏–ª–∏ –≤–≤–µ–¥–∏—Ç–µ —Å–≤–æ–π –±—Ä–µ–Ω–¥:",
                                       reply_markup=keyboard)

        # self.m = await message.answer("–í—ã–±–µ—Ä–∏—Ç–µ –±—Ä–µ–Ω–¥ –∞–≤—Ç–æ–º–æ–±–∏–ª—è:", reply_markup=keyboard)
        await state.set_state(User.STATE_CAR_BRAND)

    async def get_car_brand(self, message, state):
        search_brand = message.text
        user_data = {"car_brand": search_brand}  # –°–æ—Ö—Ä–∞–Ω—è–µ–º –Ω–∞–∑–≤–∞–Ω–∏–µ –º–∞—Ä–∫–∏ –≤ –¥–∞–Ω–Ω—ã—Ö –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è
        await state.update_data(user_data=user_data)  # –û–±–Ω–æ–≤–ª—è–µ–º –¥–∞–Ω–Ω—ã–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –≤ —Å–æ—Å—Ç–æ—è–Ω–∏–∏

        if search_brand == "‚å® –í–≤–µ–¥–∏—Ç–µ —Å–≤–æ–π –±—Ä–µ–Ω–¥":
            await message.answer("–ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –≤–≤–µ–¥–∏—Ç–µ –Ω–∞–∑–≤–∞–Ω–∏–µ –º–∞—Ä–∫–∏ —Å–≤–æ–µ–≥–æ –∞–≤—Ç–æ–º–æ–±–∏–ª—è:")
        else:
            models = await recognize_car_model(message, search_brand)

            if not models:
                await message.answer("–ú–∞—Ä–∫–∞ –Ω–µ –Ω–∞–π–¥–µ–Ω–∞, –ø–æ–ø—Ä–æ–±—É–π—Ç–µ –µ—â–µ —Ä–∞–∑")
            else:
                model_names = [model['name'] for model in models]
                keyboard = create_keyboard(model_names)
                await message.answer("–í—ã–±–µ—Ä–∏—Ç–µ –º–æ–¥–µ–ª—å –∞–≤—Ç–æ–º–æ–±–∏–ª—è –∏–∑ —Å–ø–∏—Å–∫–∞:", reply_markup=keyboard)

                response = f"–ú–æ–¥–µ–ª–∏ –∞–≤—Ç–æ–º–æ–±–∏–ª–µ–π –º–∞—Ä–∫–∏ '{search_brand}':"
                await message.answer(response, reply_markup=keyboard)
                response1 = f"–ó–∞–≥—Ä—É–∑–∏ —Ñ–æ—Ç–∫–∏"
                await message.answer(response1)
                await state.set_state(User.STATE_CAR_PHOTO)

    async def handle_photos(self, message, state):
        user_data = await state.get_data('user_data')
        photo_id = message.photo[-1].file_id

        self.new_id = str(uuid.uuid4().int)[:6]

        caption = (
            f"üõû <b>#{user_data.get('user_data').get('car_brand')}-{user_data.get('user_data').get('car_model')}</b>\n\n"
            f"üí¨<b>–¢–µ–ª–µ–≥—Ä–∞–º:</b> <span class='tg-spoiler'>@{message.from_user.username if message.from_user.username is not None else '–ø–æ –Ω–æ–º–µ—Ä—É —Ç–µ–ª–µ—Ñ–æ–Ω–∞'}</span>\n\n"
            f" {hlink('Selbie Auto. –†—ã–Ω–æ–∫ —Ç–∞—á–µ–∫ –≤ –î–ù–†', 'https://t.me/selbieauto')} | {hlink('–†–∞–∑–º–µ—Å—Ç–∏—Ç—å –∞–≤—Ç–æ', 'https://t.me/selbie_bot')} \n\n"
            f"<b>ID –æ–±—ä—è–≤–ª–µ–Ω–∏—è: #{self.new_id}</b>"
        )

        if "sent_photos" not in user_data:
            user_data["sent_photos"] = []

        user_data["sent_photos"].append({"file_id": photo_id, })
        buffered_photos.append(InputMediaPhoto(
            media=photo_id, caption=caption, parse_mode=types.ParseMode.HTML))
        if len(buffered_photos) > 1:
            for i in range(len(buffered_photos) - 1):
                buffered_photos[i].caption = None
            last_photo = buffered_photos[-1]
            last_photo.caption = caption

        keyboard = ReplyKeyboardMarkup(resize_keyboard=True).add(
            KeyboardButton("–°–ª–µ–¥—É—â–∏–π —à–∞–≥")
        )

        self.m = await message.answer("–§–æ—Ç–æ –¥–æ–±–∞–≤–ª–µ–Ω–æ", reply_markup=keyboard)

        self.db_fix = user_data

        await state.finish()

    async def add_data_to_excel(self, message):
        file_path = 'db.xlsx'

        row_data = [
            self.new_id,
            datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            self.db_fix.get('user_data').get('car_brand', ''),
            message.from_user.username if message.from_user.username is not None else '–ø–æ –Ω–æ–º–µ—Ä—É —Ç–µ–ª–µ—Ñ–æ–Ω–∞',
        ]

        # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —Å—É—â–µ—Å—Ç–≤—É–µ—Ç –ª–∏ —Ñ–∞–π–ª Excel
        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(file_path)
        else:
            workbook = openpyxl.Workbook()
        sheet = workbook.active

        # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –Ω—É–∂–Ω–æ –ª–∏ –¥–æ–±–∞–≤–∏—Ç—å –∑–∞–≥–æ–ª–æ–≤–∫–∏
        if sheet.max_row == 1:
            headers = [
                'ID', '–î–∞—Ç–∞', '–ë—Ä–µ–Ω–¥', '–¢–µ–ª–µ–≥—Ä–∞–º'
            ]
            sheet.append(headers)

        sheet.append(row_data)
        workbook.save(file_path)

    async def preview_advertisement(self, message):
        await bot.send_media_group(chat_id=message.chat.id, media=buffered_photos, disable_notification=True)

        keyboard = ReplyKeyboardMarkup(resize_keyboard=True).add(
            KeyboardButton("–û—Ç–ø—Ä–∞–≤–∏—Ç—å –≤ –∫–∞–Ω–∞–ª"),
            KeyboardButton("–û—Ç–º–µ–Ω–∏—Ç—å –∏ –∑–∞–ø–æ–ª–Ω–∏—Ç—å –∑–∞–Ω–æ–≤–æ"),
        )
        await message.reply(
            "–¢–∞–∫ –±—É–¥–µ—Ç –≤—ã–≥–ª—è–¥–µ—Ç—å –≤–∞—à–µ –æ–±—ä—è–≤–ª–µ–Ω–∏–µ. –í—ã –º–æ–∂–µ—Ç–µ –ª–∏–±–æ —Ä–∞–∑–º–µ—Å—Ç–∏—Ç—å –ª–∏–±–æ –æ—Ç–º–µ–Ω–∏—Ç—å –∏ –∑–∞–ø–æ–ª–Ω–∏—Ç—å –∑–∞–Ω–æ–≤–æ.",
            reply_markup=keyboard)

    async def send_advertisement(self, message):
        # user_id = message.from_user.id
        async with lock:
            user_id = message.from_user.id
            await self.add_data_to_excel(message)
            await bot.send_media_group(chat_id=CHANNEL_ID, media=buffered_photos, disable_notification=True)
            keyboard = create_keyboard(['–î–æ–±–∞–≤–∏—Ç—å –µ—â—ë –æ–±—ä—è–≤–ª–µ–Ω–∏–µ', '–£—Å–∫–æ—Ä–∏—Ç—å –ø—Ä–æ–¥–∞–∂—É'])
            await bot.send_message(user_id, "–û–±—ä—è–≤–ª–µ–Ω–∏–µ –æ—Ç–ø—Ä–∞–≤–ª–µ–Ω–æ –≤ –∫–∞–Ω–∞–ª!", reply_markup=keyboard)

            buffered_photos.clear()

    async def fill_again(self, message, state):
        keyboard = create_keyboard(list(dict_car_brands_and_models.keys()))
        image_path = ImageDirectory.auto_car_brand  # –ü—É—Ç—å –∫ –≤–∞—à–µ–º—É –∏–∑–æ–±—Ä–∞–∂–µ–Ω–∏—é
        with open(image_path, "rb") as image:
            self.m = await message.answer_photo(image, caption="–í—ã–±–µ—Ä–∏—Ç–µ –±—Ä–µ–Ω–¥ –∞–≤—Ç–æ–º–æ–±–∏–ª—è:", reply_markup=keyboard)
        # self.m = await message.answer("–í—ã–±–µ—Ä–∏—Ç–µ –±—Ä–µ–Ω–¥ –∞–≤—Ç–æ–º–æ–±–∏–ª—è:", reply_markup=keyboard)
        async with lock:
            buffered_photos.clear()
        await state.set_state(User.STATE_CAR_BRAND)

    async def add_more(self, message, state):
        await car_bot.restart(message, state)

    async def promotion(self, message, state):
        keyboard = create_keyboard(['–ü–µ—Ä–µ–∑–∞–≥—Ä—É–∑–∏—Ç—å –±–æ—Ç–∞'])
        await message.reply("–ß—Ç–æ–±—ã –∫—É–ø–∏—Ç—å –∑–∞–∫—Ä–µ–ø –Ω–∞–ø–∏—à–∏—Ç–µ @selbie_adv", reply_markup=keyboard)


car_bot = CarBotHandler()
bot = Bot(token=API_TOKEN)
dp = Dispatcher(bot, storage=MemoryStorage())
lock = asyncio.Lock()
buffered_photos = []


@dp.message_handler(lambda message: message.text == "–ü–µ—Ä–µ–∑–∞–≥—Ä—É–∑–∏—Ç—å –±–æ—Ç–∞", state='*')
@dp.message_handler(commands=['restart'], state='*')
async def cmd_restart(message: types.Message, state: FSMContext):
    await car_bot.restart(message, state)


@dp.message_handler(commands=["start"])
async def cmd_start(message: types.Message, state: FSMContext):
    await car_bot.start(message, state)


# support
@dp.message_handler(commands=['support'], state='*')
async def cmd_support(message: types.Message, state: FSMContext):
    await car_bot.support(message, state)


@dp.message_handler(state=User.STATE_SUPPORT_VALIDATION)
async def support_validation(message: types.Message, state: FSMContext):
    await car_bot.support_validation(message, state)


@dp.message_handler(state=User.STATE_SUPPORT_MESSAGE)
async def support_message(message: types.Message, state: FSMContext):
    await car_bot.support_message(message, state)


@dp.message_handler(state=User.STATE_SUPPORT_END)
async def support_end(message: types.Message, state: FSMContext):
    await car_bot.restart(message, state)


# end support

@dp.message_handler(state=User.STATE_CAR_BRAND)
async def get_car_brand(message: types.Message, state: FSMContext):
    await car_bot.get_car_brand(message, state)


@dp.message_handler(lambda message: message.text == '–í–≤–µ–¥–∏—Ç–µ —Å–≤–æ–π –±—Ä–µ–Ω–¥')
async def input_brand(message: types.Message):
    await message.answer("–í–≤–µ–¥–∏—Ç–µ –º–∞—Ä–∫—É –∞–≤—Ç–æ–º–æ–±–∏–ª—è –¥–ª—è –ø–æ–∏—Å–∫–∞ –º–æ–¥–µ–ª–µ–π:")

@dp.message_handler(state=User.STATE_CAR_PHOTO, content_types=['photo'])
async def handle_photos(message: types.Message, state: FSMContext):
    await car_bot.handle_photos(message, state)


@dp.message_handler(lambda message: message.text == "–°–ª–µ–¥—É—â–∏–π —à–∞–≥")
async def preview_advertisement(message: types.Message):
    await car_bot.preview_advertisement(message)


@dp.message_handler(lambda message: message.text == "–û—Ç–ø—Ä–∞–≤–∏—Ç—å –≤ –∫–∞–Ω–∞–ª")
async def send_advertisement(message: types.Message, state: FSMContext):
    await car_bot.send_advertisement(message)


@dp.message_handler(lambda message: message.text == "–û—Ç–º–µ–Ω–∏—Ç—å –∏ –∑–∞–ø–æ–ª–Ω–∏—Ç—å –∑–∞–Ω–æ–≤–æ")
async def fill_again(message: types.Message, state: FSMContext):
    await car_bot.fill_again(message, state)


@dp.message_handler(lambda message: message.text == "–î–æ–±–∞–≤–∏—Ç—å –µ—â—ë –æ–±—ä—è–≤–ª–µ–Ω–∏–µ")
async def add_more(message: types.Message, state: FSMContext):
    await car_bot.add_more(message, state)


@dp.message_handler(lambda message: message.text == "–£—Å–∫–æ—Ä–∏—Ç—å –ø—Ä–æ–¥–∞–∂—É")
async def promotion(message: types.Message, state: FSMContext):
    await car_bot.promotion(message, state)


# —Å—Ç–∞—Ä—Ç –±–æ—Ç–∞
if __name__ == '__main__':
    executor.start_polling(dp, skip_updates=True)