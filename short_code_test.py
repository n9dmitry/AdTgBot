from aiogram import Bot, Dispatcher, types, executor
from aiogram.types import InputMediaPhoto
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton
import uuid
import asyncio
import openpyxl
from config import *
from states import *
from validation import *
import json
from enumlist import *

# Загрузка JSON в начале скрипта
with open('dicts.json', 'r', encoding='utf-8') as file:
    dicts = json.load(file)

dict_car_brands_and_models = dicts.get("dict_car_brands_and_models", {})
dict_car_body_types = dicts.get("dict_car_body_types", {})
dict_car_engine_types = dicts.get("dict_car_engine_types", {})
dict_car_transmission_types = dicts.get("dict_car_transmission_types", {})
dict_car_colors = dicts.get("dict_car_colors", {})
dict_car_document_statuses = dicts.get("dict_car_document_statuses", {})
dict_car_owners = dicts.get("dict_car_owners", {})
dict_car_customs_cleared = dicts.get("dict_car_customs_cleared", {})
dict_currency = dicts.get("dict_currency", {})
dict_car_conditions = dicts.get("dict_car_conditions", {})
dict_car_mileages = dicts.get("dict_car_mileages", {})
dict_edit_buttons = dicts.get("dict_edit_buttons", {})
# Конец импорта json словарей


# Создание клавиатуры
def create_keyboard(button_texts, resize_keyboard=True):
    keyboard = ReplyKeyboardMarkup(
        resize_keyboard=resize_keyboard, row_width=2)
    buttons = [KeyboardButton(text=text) for text in button_texts]
    keyboard.add(*buttons)
    return keyboard




class CarBotHandler:
    def __init__(self):
        self.lock = asyncio.Lock()



# Начало работы бота

    async def start(self, event, state):
        image_hello_path = ImageDirectory.say_hi
        with open(image_hello_path, "rb") as image_hello:
            self.m = await event.answer_photo(image_hello,
                                     caption=f"Привет, {event.from_user.first_name}! Давай продадим твоё авто! Начнём же сбор данных!")
        # await asyncio.sleep(2)
        await self.m.delete()
        # self.m = await event.answer(f"Привет, {event.from_user.first_name}! Я бот для сбора данных. Давай начнем.")
        keyboard = create_keyboard(list(dict_car_brands_and_models.keys()))
        image_path = ImageDirectory.car_brand  # Путь к вашему изображению
        with open(image_path, "rb") as image:
            self.m = await event.answer_photo(image, caption="Выберите бренд автомобиля:", reply_markup=keyboard)
        # self.m = await event.answer("Выберите бренд автомобиля:", reply_markup=keyboard)
        await state.set_state(User.STATE_CAR_BRAND)



    async def get_car_brand(self, event, state):
        user_data = (await state.get_data()).get("user_data", {})
        await self.m.delete()

        selected_brand = event.text
        valid_brands = dict_car_brands_and_models
        if await validate_car_brand(selected_brand, valid_brands):
            user_data["car_brand"] = selected_brand
            await state.update_data(user_data=user_data)
            # await self.delete_previous_question(event)
            # await self.delete_hello(event)
            # Создаем клавиатуру
            # keyboard = create_keyboard(
            #     dict_car_brands_and_models[selected_brand])
            # image_path = ImageDirectory.car_model
            # with open(image_path, "rb") as image:
            #     self.m = await event.answer_photo(image, caption="Отлично! Выберите модель автомобиля:", reply_markup=keyboard)
            self.m = await event.answer("Отлично! Теперь фото:")
            await state.set_state(User.STATE_CAR_PHOTO)
        else:
#             await self.delete_previous_question(event)
#             await self.delete_hello(event)
            keyboard = create_keyboard(dict_car_brands_and_models.keys())
            self.m = await bot.send_message(event.from_user.id, "Пожалуйста, выберите бренд из предложенных вариантов или напишите нам если вашего бренда нет", reply_markup=keyboard)
            await state.set_state(User.STATE_CAR_BRAND)

    # async def write_data_to_excel(self, user_data):
    #     if os.path.exists('udb.xlsx'):
    #         workbook = openpyxl.load_workbook('udb.xlsx')
    #     else:
    #         workbook = openpyxl.Workbook()
    #
    #     ws = workbook.active
    #
    #     column_headers = ['Название', 'Телеграм']
    #     data_row = [user_data.get(field, '') for field in ['car_brand', 'seller_phone']]
    #
    #     # Find the next available row number
    #     row_num = ws.max_row + 1
    #
    #     # Write column headers if the file is new
    #     if row_num == 2:
    #         for col_num, header in enumerate(column_headers, start=1):
    #             ws.cell(row=1, column=col_num, value=header)
    #
    #     # Write data row
    #     for col_num, data in enumerate(data_row, start=1):
    #         ws.cell(row=row_num, column=col_num, value=data)
    #
    #     # Save the Excel file
    #     await self.save_excel_asynchronous('udb.xlsx', workbook)
    #
    # async def save_excel_asynchronous(self, filename, workbook):
    #     loop = asyncio.get_event_loop()
    #     await loop.run_in_executor(None, workbook.save, filename)

    async def write_to_excel(self, event, user_data):
        file_path = 'output.xlsx'

        # Проверяем, существует ли файл Excel
        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(file_path)
        else:
            workbook = openpyxl.Workbook()

        sheet = workbook.active

        # Проверяем, нужно ли добавить заголовки
        if sheet.max_row == 1:
            headers = [
                'Car Brand-Model', 'Year', 'Mileage (km)', 'Transmission Type',
                'Body Type', 'Engine Type', 'Engine Volume (L)', 'Power (hp)',
                'Color', 'Document Status', 'Number of Owners', 'Customs Cleared',
                'Condition', 'Additional Description', 'Price', 'Currency',
                'Location', 'Seller Name', 'Seller Phone', 'Telegram'
            ]
            sheet.append(headers)

        row_data = [
            user_data.get('user_data').get('car_brand', '') + '-' + user_data.get('user_data').get('car_model', ''),
            user_data.get('user_data').get('car_year', ''),
            user_data.get('user_data').get('car_mileage', ''),
            user_data.get('user_data').get('car_transmission_type', ''),
            user_data.get('user_data').get('car_body_type', ''),
            user_data.get('user_data').get('car_engine_type', ''),
            user_data.get('user_data').get('car_engine_volume', ''),
            user_data.get('user_data').get('car_power', ''),
            user_data.get('user_data').get('car_color', ''),
            user_data.get('user_data').get('car_document_status', ''),
            user_data.get('user_data').get('car_owners', ''),
            'Да' if user_data.get('user_data').get('car_customs_cleared') else 'Нет',
            user_data.get('user_data').get('car_condition', ''),
            user_data.get('user_data').get('car_description', ''),
            user_data.get('user_data').get('car_price', ''),
            user_data.get('user_data').get('currency', ''),
            user_data.get('user_data').get('car_location', ''),
            user_data.get('user_data').get('seller_name', ''),
            user_data.get('user_data').get('seller_phone', ''),
            event.from_user.username if event.from_user.username is not None else 'по номеру телефона',
        ]

        sheet.append(row_data)

        # Сохраняем файл
        workbook.save(file_path)

    async def handle_photos(self, event, state):
        user_id = event.from_user.id
        user_data = await state.get_data('user_data')
        photo_id = event.photo[-1].file_id

        caption = (
            f"🛞 <b>#{user_data.get('user_data').get('car_brand')}-{user_data.get('user_data').get('car_model')}</b>\n\n"
            f"   <b>-Год:</b> {user_data.get('user_data', {}).get('car_year')}\n"
            f"   <b>-Пробег (км.):</b> {user_data.get('user_data').get('car_mileage')}\n"
            f"   <b>-Тип КПП:</b> {user_data.get('user_data').get('car_transmission_type')}\n"
            f"   <b>-Кузов:</b> {user_data.get('user_data').get('car_body_type')}\n"
            f"   <b>-Тип двигателя:</b> {user_data.get('user_data').get('car_engine_type')}\n"
            f"   <b>-Объем двигателя (л.):</b> {user_data.get('user_data').get('car_engine_volume')}\n"
            f"   <b>-Мощность (л.с.):</b> {user_data.get('user_data').get('car_power')}\n"
            f"   <b>-Цвет:</b> {user_data.get('user_data').get('car_color')}\n"
            f"   <b>-Статус документов:</b> {user_data.get('user_data').get('car_document_status')}\n"
            f"   <b>-Количество владельцев:</b> {user_data.get('user_data').get('car_owners')}\n"
            f"   <b>-Растаможка:</b> {'Да' if user_data.get('user_data').get('car_customs_cleared') else 'Нет'}\n"
            f"   <b>-Состояние:</b> {user_data.get('user_data').get('car_condition')}\n\n"
            f"ℹ️ <b>Дополнительная информация:</b> {user_data.get('user_data').get('car_description')}\n\n"
            f"🔥 <b>Цена:</b> {user_data.get('user_data').get('car_price')} {user_data.get('user_data').get('currency')}\n\n"
            f"📍 <b>Местоположение:</b> {user_data.get('user_data').get('car_location')}\n"
            f"👤 <b>Продавец:</b> <span class='tg-spoiler'> {user_data.get('user_data').get('seller_name')} </span>\n"
            f"📲 <b>Телефон продавца:</b> <span class='tg-spoiler'>{user_data.get('user_data').get('seller_phone')} </span>\n"
            f"💬 <b>Телеграм:</b> <span class='tg-spoiler'>{event.from_user.username if event.from_user.username is not None else 'по номеру телефона'}</span>\n\n"
            f"ООО 'Продвижение' Авто в ДНР (link: разместить авто)"
        )

        photo_uuid = str(uuid.uuid4())

        if "sent_photos" not in user_data:
            user_data["sent_photos"] = []

        user_data["sent_photos"].append(
            {"file_id": photo_id, "uuid": photo_uuid})

        photo = InputMediaPhoto(media=photo_id, caption=caption, parse_mode=types.ParseMode.HTML)
        buffered_photos.append(photo)

        # if len(buffered_photos) > 1:
        #     for i in range(len(buffered_photos)):
        #         buffered_photos[i].caption = None

        # last_photo = buffered_photos[-1]
        # last_photo.caption = caption



        keyboard = ReplyKeyboardMarkup(resize_keyboard=True).add(
        KeyboardButton("Следущий шаг")
            )

        print(len(buffered_photos))
        await event.reply('Фото добавлено', reply_markup=keyboard)


        # # Добавляем условие для отправки сообщения при добавлении фотографий
        # if len(buffered_photos) > 1:
        #     await bot.send_media_group(chat_id=event.chat.id, media=buffered_photos, disable_notification=True)
        #     await event.answer(f"Фото были успешно добавлены. Переходим к следующему шагу.")


        await state.finish()

    async def preview_advertisement(self, event, caption):
        for i in range(len(buffered_photos)):
            buffered_photos[i].caption = None

        last_photo = buffered_photos[-1]
        last_photo.caption = self.caption

        await event.reply("Фото добавлены")
        await asyncio.sleep(2)
        await bot.send_media_group(chat_id=event.chat.id, media=buffered_photos, disable_notification=True)


        keyboard = ReplyKeyboardMarkup(resize_keyboard=True).add(
            KeyboardButton("Отправить в канал"),
            KeyboardButton("Отменить и заполнить заново"),
        )
        await event.reply(
            "Так будет выглядеть ваше объявление. Вы можете либо разместить либо отменить и заполнить заново.",
            reply_markup=keyboard)



    # async def preview_advertisement(self, event):
    #     await event.reply("Фото добавлено")
    #     await asyncio.sleep(2)
    #     await bot.send_media_group(chat_id=event.chat.id, media=buffered_photos, disable_notification=True)
    #
    #
    #     keyboard = ReplyKeyboardMarkup(resize_keyboard=True).add(
    #         KeyboardButton("Отправить в канал"),
    #         KeyboardButton("Отменить и заполнить заново"),
    #     )
    #     await event.reply(
    #         "Так будет выглядеть ваше объявление. Вы можете либо разместить либо отменить и заполнить заново.",
    #         reply_markup=keyboard)
    #
    # async def send_advertisement(self, event):
    #     user_id = event.from_user.id
    #     await self.m.delete()
    #     async with lock:
    #         await bot.send_media_group(chat_id=CHANNEL_ID, media=buffered_photos, disable_notification=True)
    #         await bot.send_message(user_id, "Объявление отправлено в канал!")
    #
    # async def fill_again(self, event, state):
    #     keyboard = create_keyboard(list(dict_car_brands_and_models.keys()))
    #     image_path = ImageDirectory.car_brand  # Путь к вашему изображению
    #     with open(image_path, "rb") as image:
    #         self.m = await event.answer_photo(image, caption="Выберите бренд автомобиля:", reply_markup=keyboard)
    #     # self.m = await event.answer("Выберите бренд автомобиля:", reply_markup=keyboard)
    #     buffered_photos.clear()
    #     await state.set_state(User.STATE_CAR_BRAND)


car_bot = CarBotHandler()
bot = Bot(token=API_TOKEN)
dp = Dispatcher(bot, storage=MemoryStorage())
lock = asyncio.Lock()
buffered_photos = []


@dp.message_handler(commands=["start"])
async def cmd_start(event: types.Message, state: FSMContext):
    await car_bot.start(event, state)


@dp.message_handler(state=User.STATE_CAR_BRAND)
async def process_brand_selection(event: types.Message, state: FSMContext):
    await car_bot.get_car_brand(event, state)

@dp.message_handler(state=User.STATE_CAR_PHOTO, content_types=['photo'])
async def handle_photos(event: types.Message, state: FSMContext):
    print('STATE:', state, event)
    await car_bot.handle_photos(event, state)

@dp.message_handler(lambda message: message.text == "Следущий шаг")
async def preview_advertisement(event: types.Message, state: FSMContext):
    await car_bot.preview_advertisement(event)


@dp.message_handler(lambda message: message.text == "Отправить в канал")
async def send_advertisement(event: types.Message, state: FSMContext):
    await car_bot.send_advertisement(event)


@dp.message_handler(lambda message: message.text == "Отменить и заполнить заново")
async def fill_again(event: types.Message, state: FSMContext):
    await car_bot.fill_again(event, state)





# старт бота
if __name__ == '__main__':
    executor.start_polling(dp, skip_updates=True)