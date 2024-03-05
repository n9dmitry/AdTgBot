from aiogram import Bot, Dispatcher, types, executor
from aiogram.types import InputMediaPhoto
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton
from aiogram.utils.markdown import hlink
import random
import datetime
import uuid
import asyncio
import openpyxl
from Levenshtein import distance
from config import *
from states import *
from validation import *
import json
from enumlist import *

# Загрузка JSON в начале скрипта
with open('dicts.json', 'r', encoding='utf-8') as file:
    dicts = json.load(file)

dict_start_brands = dicts.get("dict_start_brands", {})
dict_car_brands_and_models = dicts.get("dict_car_brands_and_models", {})
dict_car_body_types = dicts.get("dict_car_body_types", {})
dict_car_engine_types = dicts.get("dict_car_engine_types", {})
dict_car_transmission_types = dicts.get("dict_car_transmission_types", {})
dict_car_colors = dicts.get("dict_car_colors", {})
dict_car_document_statuses = dicts.get("dict_car_document_statuses", {})
dict_car_owners = dicts.get("dict_car_owners", {})
dict_car_customs_cleared = dicts.get("dict_car_customs_cleared", {})
dict_currency = dicts.get("dict_currency", {})
dict_car_conditions = dicts.get("dict_car_conditions", {})
dict_car_mileages = dicts.get("dict_car_mileages", {})
dict_edit_buttons = dicts.get("dict_edit_buttons", {})
# Конец импорта json словарей


# Создание клавиатуры
def create_keyboard(button_texts, resize_keyboard=True):
    keyboard = ReplyKeyboardMarkup(
        resize_keyboard=resize_keyboard, row_width=2)
    buttons = [KeyboardButton(text=text) for text in button_texts]
    keyboard.add(*buttons)
    return keyboard

async def recognize_car_model(event, brand_name):
    models = []
    similar_brands = []

    if brand_name.lower() in ['жигули', 'ваз', 'лада']:
        brand_name = 'Lada (ВАЗ)'

    if brand_name.lower() in ['мерседес', 'мерседес-бенц', 'mercedes-benz','mercedes', 'mercedez', 'mercedez-bens']:
        brand_name = 'Mercedes-Benz'

    with open('cars.json', encoding='utf-8') as file:
        data = json.load(file)

    found_brand = False
    for item in data:
        if 'name' in item and item['name'].lower() == brand_name.lower():
            if 'models' in item:
                models = item['models']
            found_brand = True
            break
        elif 'cyrillic-name' in item and item['cyrillic-name'].lower() == brand_name.lower():
            if 'models' in item:
                models = item['models']
            found_brand = True
            break

    if not found_brand and len(brand_name) >= 3:
        for inner_item in data:
            if 'name' in inner_item and distance(brand_name.lower(), inner_item['name'].lower()) <= 2 and \
                    inner_item['name'] not in similar_brands:
                similar_brands.append(inner_item['name'])
            elif 'cyrillic-name' in inner_item and distance(brand_name.lower(),
                                                            inner_item['cyrillic-name'].lower()) <= 2 and \
                    inner_item['name'] not in similar_brands:
                similar_brands.append(inner_item['name'])

        if similar_brands:
            response_message = "Похожие бренды:\n" + "\n".join(similar_brands)
            await event.answer(response_message)

    return models




class CarBotHandler:
    def __init__(self):
        self.lock = asyncio.Lock()


# Команды

    async def restart(self, message, state):
        await state.finish()
        await message.answer("Бот перезапущен.")
        await self.start(message, state)
    async def support(self, message, state):
        await state.finish()
        self.secret_number = str(random.randint(100, 999))

        await message.answer(f"Нашли баг? Давайте отправим сообщение разработчикам! "
                             f"Но перед этим введите проверку. Докажите что вы не робот. Напишите число {self.secret_number}:")
        await state.set_state(User.STATE_SUPPORT_VALIDATION)
    async def support_validation(self, message, state):
        if message.text.isdigit() and message.text == self.secret_number:
            await message.reply(f"Проверка пройдена успешно!")
            await asyncio.sleep(1)
            await message.answer(f"Опишите техническую проблему в деталях для разработчиков: ")
            await state.set_state(User.STATE_SUPPORT_MESSAGE)
        else:
            await message.answer(f"Попробуйте ещё раз!")
            await asyncio.sleep(1)
            await cmd_support(message, state)
    async def support_message(self, message: types.Message, state):
        # Получаем текущую дату и время
        current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        # Формируем строку для записи в файл
        message_to_write = f"""
        Дата: {current_time}
        Имя: {message.from_user.full_name}
        Telegram @{message.from_user.username or message.from_user.id} 
      
        Сообщение: {message.text}
        ...
            """

        # Открываем файл для записи и записываем сообщение
        with open("support.txt", "a") as file:
            file.write(message_to_write)
        keyboard = create_keyboard(['Перезагрузить бота'])
        await message.reply("Спасибо за ваше сообщение! Мы рассмотрим вашу проблему!", reply_markup=keyboard)
        await state.set_state(User.STATE_SUPPORT_END)
    async def support_end(selfself, message, state):
        if message.text == 'Перезагрузить бота':
            await cmd_restart(message, state)
        await state.finish()

# Начало работы бота
#     async def start(self, message, state):
#         image_hello_path = ImageDirectory.auto_say_hi
#         with open(image_hello_path, "rb") as image_hello:
#             self.m = await message.answer_photo(image_hello,
#                                      caption=f"Привет, {message.from_user.first_name}! Давай продадим твоё авто! Начнём же сбор данных!")
#         await asyncio.sleep(0)
#         # self.m = await message.answer(f"Привет, {message.from_user.first_name}! Я бот для сбора данных. Давай начнем.")
#         keyboard = create_keyboard(list(dict_car_brands_and_models.keys()))
#         image_path = ImageDirectory.auto_car_brand  # Путь к вашему изображению
#         with open(image_path, "rb") as image:
#             self.m = await message.answer_photo(image, caption="Выберите бренд автомобиля:", reply_markup=keyboard)
#         # self.m = await message.answer("Выберите бренд автомобиля:", reply_markup=keyboard)
#         await state.set_state(User.STATE_CAR_BRAND)

    async def start(self, message, state):
        image_hello_path = ImageDirectory.auto_say_hi
        with open(image_hello_path, "rb") as image_hello:
            self.m = await message.answer_photo(image_hello,
                                     caption=f"Привет, {message.from_user.first_name}! Давай продадим твоё авто! Начнём же сбор данных!")
        await asyncio.sleep(0)

        keyboard = ReplyKeyboardMarkup(resize_keyboard=True)
        target_brands = [
            "Lada", "Kia", "Hyundai", "Toyota", "Volkswagen", "Nissan", "Renault", "Skoda", "Ford", "Mercedes-Benz"
        ]
        with open('cars.json', encoding='utf-8') as file:
            data = json.load(file)

            for item in data:
                if item['name'] in target_brands:
                    keyboard.add(KeyboardButton(text=item['name']))
        keyboard.add(KeyboardButton(text="Введите свой бренд"))
        image_path = ImageDirectory.auto_car_brand  # Путь к вашему изображению
        with open(image_path, "rb") as image:
            await message.answer_photo(image, caption="Выберите одну из кнопок ниже или введите свой бренд:", reply_markup=keyboard)



        # self.m = await message.answer("Выберите бренд автомобиля:", reply_markup=keyboard)
        await state.set_state(User.STATE_CAR_BRAND)


    async def get_car_brand(self, message, state):
        search_brand = message.text
        models = await recognize_car_model(message, search_brand)

        if models:
            keyboard = ReplyKeyboardMarkup(resize_keyboard=True)
            added_models = set()  # Множество для хранения добавленных моделей

            for model in models:
                model_name = model['name']
                if model_name not in added_models:
                    button_text = f"{model_name}"
                    keyboard.add(KeyboardButton(text=button_text))
                    added_models.add(model_name)

            response = f"Модели автомобилей марки '{search_brand}':"
            await message.answer(response, reply_markup=keyboard)
        else:
            await message.answer(f"Модели автомобилей марки '{search_brand}' не найдены.")

    # async def get_car_brand(self, message, state):
    #     user_data = (await state.get_data()).get("user_data", {})
    #     selected_brand = message.text
    #     valid_brands = dict_car_brands_and_models
    #     if await validate_car_brand(selected_brand, valid_brands):
    #         user_data["car_brand"] = selected_brand
    #         await state.update_data(user_data=user_data)
    #         # Создаем клавиатуру
    #         keyboard = create_keyboard(
    #             dict_car_brands_and_models[selected_brand])
    #         image_path = ImageDirectory.auto_car_model
    #         with open(image_path, "rb") as image:
    #             self.m = await message.answer_photo(image, caption="Отлично! Выберите модель автомобиля:", reply_markup=keyboard)
    #         # self.m = await message.answer("Отлично! Выберите модель автомобиля:", reply_markup=keyboard)
    #         await state.set_state(User.STATE_CAR_MODEL)
    #     else:
    #         keyboard = create_keyboard(dict_car_brands_and_models.keys())
    #         self.m = await bot.send_message(message.from_user.id, "Пожалуйста, выберите бренд из предложенных вариантов или напишите нам если вашего бренда нет", reply_markup=keyboard)
    #         await state.set_state(User.STATE_CAR_BRAND)

    async def get_car_model(self, message, state):
        user_data = (await state.get_data()).get("user_data", {})
        car_brand = user_data.get("car_brand", "")
        valid_models = dict_car_brands_and_models.get(car_brand, [])

        if await validate_car_model(message.text, valid_models):
            user_data["car_model"] = message.text
            await state.update_data(user_data=user_data)
            image_path = ImageDirectory.auto_car_year
            with open(image_path, "rb") as image:
                self.m = await message.answer_photo(image, caption="Какой год выпуска у автомобиля? (напишите)")
            # await message.answer("Какой год выпуска у автомобиля? (напишите)")
            await state.set_state(User.STATE_CAR_YEAR)
        else:
            keyboard = create_keyboard(valid_models)
            self.m = await bot.send_message(message.from_user.id, "Пожалуйста, выберите модель из предложенных вариантов.",
                                   reply_markup=keyboard)
            await state.set_state(User.STATE_CAR_MODEL)


    async def get_car_year(self, message, state):
        user_data = (await state.get_data()).get("user_data", {})

        if await validate_year(message.text):
            user_data["car_year"] = message.text
            keyboard = create_keyboard(dict_car_body_types)
            await state.update_data(user_data=user_data)
            image_path = ImageDirectory.auto_car_body_type
            with open(image_path, "rb") as image:
                self.m = await message.answer_photo(image, caption="Отлично! Какой тип кузова у автомобиля?", reply_markup=keyboard)
            # self.m = await message.answer("Отлично! Какой тип кузова у автомобиля?", reply_markup=keyboard)
            await state.set_state(User.STATE_CAR_BODY_TYPE)
        else:
            self.m = await message.answer("Пожалуйста, введите год в формате YYYY (например, 1990 или 2024)")
            await state.set_state(User.STATE_CAR_YEAR)

    async def get_car_body_type(self, message, state):
        user_data = (await state.get_data()).get("user_data", {})
        if await validate_button_input(message.text, dict_car_body_types):
            user_data["car_body_type"] = message.text
            keyboard = create_keyboard(dict_car_engine_types)
            await state.update_data(user_data=user_data)
            image_path = ImageDirectory.auto_car_engine_type
            with open(image_path, "rb") as image:
                self.m = await message.answer_photo(image, caption="Отлично! Какой тип двигателя у автомобиля?", reply_markup=keyboard)
            # self.m = await message.answer("Отлично! Какой тип двигателя у автомобиля?", reply_markup=keyboard)
            await state.set_state(User.STATE_CAR_ENGINE_TYPE)
        else:
            keyboard = create_keyboard(dict_car_body_types)
            self.m = await message.answer("Пожалуйста, выберите корректный тип кузова.", reply_markup=keyboard)
            await state.set_state(User.STATE_CAR_BODY_TYPE)

    async def get_car_engine_type(self, message, state):
        user_data = (await state.get_data()).get("user_data", {})
        if await validate_button_input(message.text, dict_car_engine_types):
            user_data["car_engine_type"] = message.text
            # Добавляем кнопки на основе словаря
            await state.update_data(user_data=user_data)
            image_path = ImageDirectory.auto_car_engine_volume
            with open(image_path, "rb") as image:
                self.m = self.m = await message.answer_photo(image, caption="Хорошо! Какой объем двигателя у автомобиля (л.)? (напишите через точку: например 1.6)")
            # self.m = await message.answer("Хорошо! Какой объем двигателя у автомобиля (л.)? (напишите через точку: например 1.6)")
            await state.set_state(User.STATE_CAR_ENGINE_VOLUME)
        else:
            keyboard = create_keyboard(dict_car_engine_types)
            self.m = await message.answer("Пожалуйста, выберите корректный тип двигателя.", reply_markup=keyboard)
            await state.set_state(User.STATE_CAR_ENGINE_TYPE)

    async def get_car_engine_volume(self, message, state):
        user_data = (await state.get_data()).get("user_data", {})
        try:
            if "," in message.text:
                message.text = message.text.replace(',', '.')
            message.text = float(message.text)

            if await validate_engine_volume(message.text) and 0.2 <= message.text <= 10.0:
                user_data["car_engine_volume"] = message.text

                # Добавляем кнопки на основе словаря

                await state.update_data(user_data=user_data)
                image_path = ImageDirectory.auto_car_power
                with open(image_path, "rb") as image:
                    self.m = await message.answer_photo(image,
                                             caption="Отлично! Укажите мощность двигателя автомобиля от 50 до 1000 (л.с.). (напишите)")
                # self.m = await message.answer("Отлично! Укажите мощность двигателя автомобиля от 50 до 1000 (л.с.). (напишите)")
                await state.set_state(User.STATE_CAR_POWER)
            else:
                await message.answer(
                    "Пожалуйста, корректный объем двигателя (в пределах от 0.2 до 10.0 литров) через точку или целым числом(!).")
                await state.set_state(User.STATE_CAR_ENGINE_VOLUME)


        except ValueError:
            # Если не удалось преобразовать введенный текст в число
            self.m = await message.answer(
                "Пожалуйста, корректный объем двигателя (в пределах от 0.2 до 10.0 литров) через точку или целым числом(!).")
            await state.set_state(User.STATE_CAR_ENGINE_VOLUME)

    async def get_car_power(self, message, state):
        user_data = (await state.get_data()).get("user_data", {})
        if await validate_car_power(message.text):
            user_data["car_power"] = message.text
            keyboard = create_keyboard(dict_car_transmission_types)

            await state.update_data(user_data=user_data)
            image_path = ImageDirectory.auto_car_transmission_type
            with open(image_path, "rb") as image:
                self.m = await message.answer_photo(image, caption="Отлично! Какой тип коробки передач используется в автомобиле?", reply_markup=keyboard)
            # await message.answer("Отлично! Какой тип коробки передач используется в автомобиле?", reply_markup=keyboard)
            await state.set_state(User.STATE_CAR_TRANSMISSION_TYPE)
        else:
            self.m = await message.answer("Пожалуйста, введите корректную мощность двигателя (в пределах от 50 до 1000 л.с.).")
            await state.set_state(User.STATE_CAR_POWER)

    async def get_car_transmission_type(self, message, state):
        user_data = (await state.get_data()).get("user_data", {})
        if await validate_button_input(message.text, dict_car_transmission_types):
            user_data["car_transmission_type"] = message.text
            keyboard = create_keyboard(dict_car_colors)
            await state.update_data(user_data=user_data)
            image_path = ImageDirectory.auto_car_color
            with open(image_path, "rb") as image:
                self.m = await message.answer_photo(image, caption="Какого цвета автомобиль?", reply_markup=keyboard)
            # self.m = await message.answer("Какого цвета автомобиль?", reply_markup=keyboard)
            await state.set_state(User.STATE_CAR_COLOR)
        else:
            keyboard = create_keyboard(dict_car_transmission_types)
            self.m = await message.answer("Пожалуйста, выберите корректный тип трансмиссии.", reply_markup=keyboard)
            await state.set_state(User.STATE_CAR_TRANSMISSION_TYPE)

    async def get_car_color(self, message, state):
        user_data = (await state.get_data()).get("user_data", {})
        if await validate_button_input(message.text, dict_car_colors):
            user_data["car_color"] = message.text
            keyboard = create_keyboard(dict_car_mileages)
            await state.update_data(user_data=user_data)
            image_path = ImageDirectory.auto_car_mileage
            with open(image_path, "rb") as image:
                self.m = await message.answer_photo(image, caption="Каков пробег автомобиля(км.)? (если новый, выберите 'Новый')", reply_markup=keyboard)
            # self.m = await message.answer("Каков пробег автомобиля(км.)? (если новый, выберите 'Новый')", reply_markup=keyboard)
            await state.set_state(User.STATE_CAR_MILEAGE)
        else:
            keyboard = create_keyboard(dict_car_colors)
            self.m = await message.answer("Пожалуйста, выберите корректный цвет автомобиля.", reply_markup=keyboard)
            await state.set_state(User.STATE_CAR_COLOR)

    async def get_car_mileage(self, message, state):
        user_data = (await state.get_data()).get("user_data", {})
        if await validate_car_mileage(message.text):
            user_data["car_mileage"] = message.text
            keyboard = create_keyboard(dict_car_document_statuses)
            await state.update_data(user_data=user_data)
            image_path = ImageDirectory.auto_car_document_status
            with open(image_path, "rb") as image:
                self.m = await message.answer_photo(image, caption="Каков статус документов у автомобиля ?", reply_markup=keyboard)
            # self.m = await message.answer("Каков статус документов у автомобиля ?", reply_markup=keyboard)
            await state.set_state(User.STATE_CAR_DOCUMENT_STATUS)
        else:
            keyboard = create_keyboard(dict_car_mileages)
            self.m = await message.answer("Пожалуйста, введите корректное значение пробега.", reply_markup=keyboard)
            await state.set_state(User.STATE_CAR_MILEAGE)

    async def get_car_document_status(self, message, state):
        user_data = (await state.get_data()).get("user_data", {})
        if await validate_button_input(message.text, dict_car_document_statuses):

            user_data["car_document_status"] = message.text
            keyboard = create_keyboard(dict_car_owners)
            await state.update_data(user_data=user_data)
            image_path = ImageDirectory.auto_car_owners
            with open(image_path, "rb") as image:
                self.m = await message.answer_photo(image, caption="Сколько владельцев у автомобиля?", reply_markup=keyboard)
            # self.m = await message.answer("Сколько владельцев у автомобиля?", reply_markup=keyboard)
            await state.set_state(User.STATE_CAR_OWNERS)
        else:
            keyboard = create_keyboard(dict_car_document_statuses)
            self.m = await message.answer("Пожалуйста, выберите корректный статус документов автомобиля.", reply_markup=keyboard)
            await state.set_state(User.STATE_CAR_DOCUMENT_STATUS)

    async def get_car_owners(self, message, state):
        user_data = (await state.get_data()).get("user_data", {})
        if await validate_button_input(message.text, dict_car_owners):
            user_data["car_owners"] = message.text
            keyboard = create_keyboard(dict_car_customs_cleared)
            await state.update_data(user_data=user_data)
            image_path = ImageDirectory.auto_car_customs_cleared
            with open(image_path, "rb") as image:
                self.m = await message.answer_photo(image, caption="Растаможен ли автомобиль?", reply_markup=keyboard)
            # self.m = await message.answer("Растаможен ли автомобиль?", reply_markup=keyboard)
            await state.set_state(User.STATE_CAR_CUSTOMS_CLEARED)
        else:
            keyboard = create_keyboard(dict_car_owners)
            self.m = await message.answer("Пожалуйста, выберите корректное количество владельцев автомобиля.", reply_markup=keyboard)
            await state.set_state(User.STATE_CAR_OWNERS)

    async def get_car_customs_cleared(self, message, state):
        user_data = (await state.get_data()).get("user_data", {})
        if await validate_button_input(message.text, dict_car_customs_cleared):
            user_data["car_customs_cleared"] = message.text
            keyboard = create_keyboard(dict_car_conditions)
            await state.update_data(user_data=user_data)
            image_path = ImageDirectory.auto_car_condition
            with open(image_path, "rb") as image:
                self.m = await message.answer_photo(image, caption="Выберите состояние автомобиля:", reply_markup=keyboard)
            # self.m = await message.answer("Выберите состояние автомобиля:", reply_markup=keyboard)
            await state.set_state(User.STATE_CAR_CONDITION)
        else:
            keyboard = create_keyboard(dict_car_customs_cleared)
            self.m = await message.answer("Пожалуйста, выберите корректный статус растаможки автомобиля.", reply_markup=keyboard)
            await state.set_state(User.STATE_CAR_CUSTOMS_CLEARED)

    async def get_car_condition(self, message, state):
        user_data = (await state.get_data()).get("user_data", {})
        if await validate_button_input(message.text, dict_car_conditions):
            user_data["car_condition"] = message.text
            await state.update_data(user_data=user_data)
            image_path = ImageDirectory.auto_car_description
            with open(image_path, "rb") as image:
                self.m = await message.answer_photo(image, caption="Описание автомобиля. (напишите до 350 символов)")
            # self.m = await message.answer("Описание автомобиля. (напишите)")
            await state.set_state(User.STATE_CAR_DESCRIPTION)
        else:
            keyboard = create_keyboard(dict_car_conditions)
            self.m = await message.answer("Пожалуйста, выберите корректное состояние автомобиля.", reply_markup=keyboard)
            await state.set_state(User.STATE_CAR_CONDITION)

    async def get_car_description(self, message, state):
        user_data = (await state.get_data()).get("user_data", {})
        if await validate_length_text(message):
            if await validate_car_description(message.text):
                user_data["car_description"] = message.text
                keyboard = create_keyboard(dict_currency)
                await state.update_data(user_data=user_data)
                image_path = ImageDirectory.auto_car_currency
                with open(image_path, "rb") as image:
                    self.m = await message.answer_photo(image, caption="Выберите валюту:", reply_markup=keyboard)
                # self.m = await message.answer("Выберите валюту:", reply_markup=keyboard)
                await state.set_state(User.STATE_SELECT_CURRENCY)
            else:
                self.m = await message.answer("Пожалуйста, введите корректное описание.")
                await state.set_state(User.STATE_CAR_DESCRIPTION)
        else:
            self.m = await message.answer("Ваше описание сильно большое. Напишите до ~350 символов:")
            await state.set_state(User.STATE_CAR_DESCRIPTION)

    async def select_currency(self, message, state):
        user_data = (await state.get_data()).get("user_data", {})
        if await validate_button_input(message.text, dict_currency):
            user_data["currency"] = message.text
            await state.update_data(user_data=user_data)
            image_path = ImageDirectory.auto_car_price
            with open(image_path, "rb") as image:
                self.m = await message.answer_photo(image, caption="Цена автомобиля?")
            # self.m = await message.answer("Цена автомобиля?")
            await state.set_state(User.STATE_CAR_PRICE)
        else:
            keyboard = create_keyboard(dict_currency)
            self.m = await message.answer("Пожалуйста, выберите корректную валюту.", reply_markup=keyboard)
            await state.set_state(User.STATE_SELECT_CURRENCY)

    async def get_car_price(self, message, state):
        user_data = (await state.get_data()).get("user_data", {})
        if await validate_car_price(message.text):
            user_data["car_price"] = message.text
            await state.update_data(user_data=user_data)
            image_path = ImageDirectory.auto_car_location
            with open(image_path, "rb") as image:
                self.m = await message.answer_photo(image, caption="Прекрасно! Где находится автомобиль? Город/пункт. (напишите)")
            # self.m = await message.answer("Прекрасно! Где находится автомобиль? Город/пункт. (напишите)")
            await state.set_state(User.STATE_CAR_LOCATION)
        else:
            self.m = await message.answer("Пожалуйста, введите корректную цену.")
            await state.set_state(User.STATE_CAR_PRICE)

    async def get_car_location(self, message, state):
        user_data = (await state.get_data()).get("user_data", {})
        if await validate_car_location(message.text):
            user_data["car_location"] = message.text
            await state.update_data(user_data=user_data)
            image_path = ImageDirectory.auto_seller_name
            with open(image_path, "rb") as image:
                self.m = await message.answer_photo(image, caption="Прекрасно! Укажите имя продавца. (напишите)")
            # self.m = await message.answer("Прекрасно! Укажите имя продавца. (напишите)")
            await state.set_state(User.STATE_SELLER_NAME)
        else:
            self.m = await message.answer("Пожалуйста, введите корректные данные.")
            await state.set_state(User.STATE_CAR_LOCATION)

    async def get_seller_name(self, message, state):
        user_data = (await state.get_data()).get("user_data", {})
        if await validate_name(message.text) is True:
            user_data["seller_name"] = message.text
            await state.update_data(user_data=user_data)
            image_path = ImageDirectory.auto_seller_phone
            with open(image_path, "rb") as image:
                self.m = await message.answer_photo(image, caption="Отлично! Какой телефонный номер у продавца? (напишите в формате +7XXXNNNXXNN или 8XXXNNNXXNN)")
            # self.m = await message.answer("Отлично! Какой телефонный номер у продавца? (напишите в формате +7XXXNNNXXNN)")
            await state.set_state(User.STATE_SELLER_PHONE)
        else:
            self.m = await message.answer("Пожалуйста, введите корректное имя.")
            await state.set_state(User.STATE_SELLER_NAME)

    async def get_seller_phone(self, message, state):

        user_data = (await state.get_data()).get("user_data", {})
        if await validate_phone_number(message.text) is True:
            message.text = '+7' + message.text[1:] if message.text.startswith('8') else message.text
            user_data["seller_phone"] = message.text
            await state.update_data(user_data=user_data)
            print(user_data)
            if await validate_final_length(message, state, user_data):
                image_path = ImageDirectory.auto_car_photos
                with open(image_path, "rb") as image:
                    self.m = await message.answer_photo(image, caption="Добавьте фотографии авто до 10 штук (За один раз!)")
                # self.m = await message.answer("Добавьте фотографии авто")
                await state.set_state(User.STATE_CAR_PHOTO)
            else:
                await message.reply(f"Ваше сообщение получилось сильно большим! \nПерезагрузите бота и напишите объявление заново.")

        else:
            self.m = await message.answer("Пожалуйста, введите корректный номер в формате +7XXXNNNXXNN.")
            await state.set_state(User.STATE_SELLER_PHONE)

    async def handle_photos(self, message, state):
        user_data = await state.get_data('user_data')
        photo_id = message.photo[-1].file_id


        self.new_id = str(uuid.uuid4().int)[:6]

        caption = (
            f"🛞 <b>#{user_data.get('user_data').get('car_brand')}-{user_data.get('user_data').get('car_model')}</b>\n\n"
            f"   <b>-Год:</b> {user_data.get('user_data', {}).get('car_year')}\n"
            f"   <b>-Пробег (км.):</b> {user_data.get('user_data').get('car_mileage')}\n"
            f"   <b>-Тип КПП:</b> {user_data.get('user_data').get('car_transmission_type')}\n"
            f"   <b>-Кузов:</b> {user_data.get('user_data').get('car_body_type')}\n"
            f"   <b>-Тип двигателя:</b> {user_data.get('user_data').get('car_engine_type')}\n"
            f"   <b>-Объем двигателя (л.):</b> {user_data.get('user_data').get('car_engine_volume')}\n"
            f"   <b>-Мощность (л.с.):</b> {user_data.get('user_data').get('car_power')}\n"
            f"   <b>-Цвет:</b> {user_data.get('user_data').get('car_color')}\n"
            f"   <b>-Статус документов:</b> {user_data.get('user_data').get('car_document_status')}\n"
            f"   <b>-Количество владельцев:</b> {user_data.get('user_data').get('car_owners')}\n"
            f"   <b>-Растаможка:</b> {'Да' if user_data.get('user_data').get('car_customs_cleared') else 'Нет'}\n"
            f"   <b>-Состояние:</b> {user_data.get('user_data').get('car_condition')}\n\n"
            f"ℹ️<b>Дополнительная информация:</b> {user_data.get('user_data').get('car_description')}\n\n"
            f"🔥<b>Цена:</b> {user_data.get('user_data').get('car_price')} {user_data.get('user_data').get('currency')}\n\n"
            f"📍<b>Местоположение:</b> {user_data.get('user_data').get('car_location')}\n"
            f"👤<b>Продавец:</b> <span class='tg-spoiler'> {user_data.get('user_data').get('seller_name')} </span>\n"
            f"📲<b>Телефон продавца:</b> <span class='tg-spoiler'>{user_data.get('user_data').get('seller_phone')} </span>\n"
            f"💬<b>Телеграм:</b> <span class='tg-spoiler'>@{message.from_user.username if message.from_user.username is not None else 'по номеру телефона'}</span>\n\n"
            f" {hlink('Selbie Auto. Рынок тачек в ДНР', 'https://t.me/selbieauto')} | {hlink('Разместить авто', 'https://t.me/selbie_bot')} \n\n"
            f"<b>ID объявления: #{self.new_id}</b>"
        )

        if "sent_photos" not in user_data:
            user_data["sent_photos"] = []

        user_data["sent_photos"].append({"file_id": photo_id,})
        buffered_photos.append(InputMediaPhoto(
            media=photo_id, caption=caption, parse_mode=types.ParseMode.HTML))
        if len(buffered_photos) > 1:
            for i in range(len(buffered_photos) - 1):
                buffered_photos[i].caption = None
            last_photo = buffered_photos[-1]
            last_photo.caption = caption


        keyboard = ReplyKeyboardMarkup(resize_keyboard=True).add(
            KeyboardButton("Следущий шаг")
        )

        self.m = await message.answer("Фото добавлено", reply_markup=keyboard)

        self.db_fix = user_data

        await state.finish()

    async def add_data_to_excel(self, message):
        file_path = 'db.xlsx'


        row_data = [
            self.new_id,
            datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            self.db_fix.get('user_data').get('car_brand', ''),
            self.db_fix.get('user_data').get('car_model', ''),
            self.db_fix.get('user_data').get('car_year', ''),
            self.db_fix.get('user_data').get('car_mileage', ''),
            self.db_fix.get('user_data').get('car_transmission_type', ''),
            self.db_fix.get('user_data').get('car_body_type', ''),
            self.db_fix.get('user_data').get('car_engine_type', ''),
            self.db_fix.get('user_data').get('car_engine_volume', ''),
            self.db_fix.get('user_data').get('car_power', ''),
            self.db_fix.get('user_data').get('car_color', ''),
            self.db_fix.get('user_data').get('car_document_status', ''),
            self.db_fix.get('user_data').get('car_owners', ''),
            self.db_fix.get('user_data').get('car_customs_cleared'),
            self.db_fix.get('user_data').get('car_condition', ''),
            self.db_fix.get('user_data').get('car_description', ''),
            self.db_fix.get('user_data').get('car_price', ''),
            self.db_fix.get('user_data').get('currency', ''),
            self.db_fix.get('user_data').get('car_location', ''),
            self.db_fix.get('user_data').get('seller_name', ''),
            self.db_fix.get('user_data').get('seller_phone', ''),
            message.from_user.username if message.from_user.username is not None else 'по номеру телефона',
        ]

        # Проверяем, существует ли файл Excel
        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(file_path)
        else:
            workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Проверяем, нужно ли добавить заголовки
        if sheet.max_row == 1:
            headers = [
                'ID','Дата', 'Бренд', 'Модель', 'Год', 'Пробег (км)', 'Тип трансмиссии',
                'Тип кузова', 'Тип двигателя', 'Объем двигателя (л)', 'Мощность (л.с.)',
                'Цвет', 'Статус документа', 'Количество владельцев', 'Растаможен',
                'Состояние', 'Дополнительное описание', 'Цена', 'Валюта',
                'Местоположение', 'Имя продавца', 'Телефон продавца', 'Телеграм'
            ]
            sheet.append(headers)

        sheet.append(row_data)
        workbook.save(file_path)

    async def preview_advertisement(self, message):
        await bot.send_media_group(chat_id=message.chat.id, media=buffered_photos, disable_notification=True)

        keyboard = ReplyKeyboardMarkup(resize_keyboard=True).add(
            KeyboardButton("Отправить в канал"),
            KeyboardButton("Отменить и заполнить заново"),
        )
        await message.reply("Так будет выглядеть ваше объявление. Вы можете либо разместить либо отменить и заполнить заново.", reply_markup=keyboard)
    async def send_advertisement(self, message):
        # user_id = message.from_user.id
        async with lock:
            user_id = message.from_user.id
            await self.add_data_to_excel(message)
            await bot.send_media_group(chat_id=CHANNEL_ID, media=buffered_photos, disable_notification=True)
            keyboard = create_keyboard(['Добавить ещё объявление', 'Ускорить продажу'])
            await bot.send_message(user_id, "Объявление отправлено в канал!", reply_markup=keyboard)

            buffered_photos.clear()
    async def fill_again(self, message, state):
        keyboard = create_keyboard(list(dict_car_brands_and_models.keys()))
        image_path = ImageDirectory.auto_car_brand # Путь к вашему изображению
        with open(image_path, "rb") as image:
            self.m = await message.answer_photo(image, caption="Выберите бренд автомобиля:", reply_markup=keyboard)
        # self.m = await message.answer("Выберите бренд автомобиля:", reply_markup=keyboard)
        async with lock:
            buffered_photos.clear()
        await state.set_state(User.STATE_CAR_BRAND)
    async def add_more(self, message, state):
        await car_bot.restart(message, state)
    async def promotion(self, message, state):
        keyboard = create_keyboard(['Перезагрузить бота'])
        await message.reply("Чтобы купить закреп напишите @selbie_adv", reply_markup=keyboard)


car_bot = CarBotHandler()
bot = Bot(token=API_TOKEN)
dp = Dispatcher(bot, storage=MemoryStorage())
lock = asyncio.Lock()
buffered_photos = []


@dp.message_handler(lambda message: message.text == "Перезагрузить бота", state='*')
@dp.message_handler(commands=['restart'], state='*')
async def cmd_restart(message: types.Message, state: FSMContext):
    await car_bot.restart(message, state)


@dp.message_handler(commands=["start"])
async def cmd_start(message: types.Message, state: FSMContext):
    await car_bot.start(message, state)

#support
@dp.message_handler(commands=['support'], state='*')
async def cmd_support(message: types.Message, state: FSMContext):
    await car_bot.support(message, state)

@dp.message_handler(state=User.STATE_SUPPORT_VALIDATION)
async def support_validation(message: types.Message, state: FSMContext):
    await car_bot.support_validation(message, state)

@dp.message_handler(state=User.STATE_SUPPORT_MESSAGE)
async def support_message(message: types.Message, state: FSMContext):
    await car_bot.support_message(message, state)

@dp.message_handler(state=User.STATE_SUPPORT_END)
async def support_end(message: types.Message, state: FSMContext):
    await car_bot.restart(message, state)
# end support

@dp.message_handler(state=User.STATE_CAR_BRAND)
async def process_brand_selection(message: types.Message, state: FSMContext):
    await car_bot.get_car_brand(message, state)

@dp.message_handler(lambda message: message.text == "Ввести свою марку авто")
async def input_brand(message: types.Message):
    await message.answer("Введите марку автомобиля для поиска моделей:")

@dp.message_handler(state=User.STATE_CAR_MODEL)
async def process_model(message: types.Message, state: FSMContext):
    await car_bot.get_car_model(message, state)


@dp.message_handler(state=User.STATE_CAR_YEAR)
async def get_car_year_handler(message: types.Message, state: FSMContext):
    await car_bot.get_car_year(message, state)


@dp.message_handler(state=User.STATE_CAR_BODY_TYPE)
async def get_car_body_type(message: types.Message, state: FSMContext):
    await car_bot.get_car_body_type(message, state)


@dp.message_handler(state=User.STATE_CAR_ENGINE_TYPE)
async def get_car_engine_type(message: types.Message, state: FSMContext):
    await car_bot.get_car_engine_type(message, state)


@dp.message_handler(state=User.STATE_CAR_ENGINE_VOLUME)
async def get_car_engine_volume(message: types.Message, state: FSMContext):
    await car_bot.get_car_engine_volume(message, state)


@dp.message_handler(state=User.STATE_CAR_POWER)
async def get_car_power(message: types.Message, state: FSMContext):
    await car_bot.get_car_power(message, state)


@dp.message_handler(state=User.STATE_CAR_TRANSMISSION_TYPE)
async def get_car_transmission_type(message: types.Message, state: FSMContext):
    await car_bot.get_car_transmission_type(message, state)


@dp.message_handler(state=User.STATE_CAR_COLOR)
async def get_car_color(message: types.Message, state: FSMContext):
    await car_bot.get_car_color(message, state)


@dp.message_handler(state=User.STATE_CAR_MILEAGE)
async def get_car_mileage(message: types.Message, state: FSMContext):
    await car_bot.get_car_mileage(message, state)


@dp.message_handler(state=User.STATE_CAR_DOCUMENT_STATUS)
async def get_car_document_status(message: types.Message, state: FSMContext):
    await car_bot.get_car_document_status(message, state)


@dp.message_handler(state=User.STATE_CAR_OWNERS)
async def get_car_owners(message: types.Message, state: FSMContext):
    await car_bot.get_car_owners(message, state)


@dp.message_handler(state=User.STATE_CAR_CUSTOMS_CLEARED)
async def get_car_customs_cleared(message: types.Message, state: FSMContext):
    await car_bot.get_car_customs_cleared(message, state)


@dp.message_handler(state=User.STATE_CAR_CONDITION)
async def get_car_condition(message: types.Message, state: FSMContext):
    await car_bot.get_car_condition(message, state)


@dp.message_handler(state=User.STATE_CAR_DESCRIPTION)
async def get_car_description(message: types.Message, state: FSMContext):
    await car_bot.get_car_description(message, state)


@dp.message_handler(state=User.STATE_SELECT_CURRENCY)
async def select_currency(message: types.Message, state: FSMContext):
    await car_bot.select_currency(message, state)


@dp.message_handler(state=User.STATE_CAR_PRICE)
async def get_car_price(message: types.Message, state: FSMContext):
    await car_bot.get_car_price(message, state)


@dp.message_handler(state=User.STATE_CAR_LOCATION)
async def get_car_location_handler(message: types.Message, state: FSMContext):
    await car_bot.get_car_location(message, state)


@dp.message_handler(state=User.STATE_SELLER_NAME)
async def get_seller_name_handler(message: types.Message, state: FSMContext):
    await car_bot.get_seller_name(message, state)


@dp.message_handler(state=User.STATE_SELLER_PHONE)
async def get_seller_phone_handler(message: types.Message, state: FSMContext):
    await car_bot.get_seller_phone(message, state)


@dp.message_handler(state=User.STATE_CAR_PHOTO, content_types=['photo'])
async def handle_photos(message: types.Message, state: FSMContext):
    await car_bot.handle_photos(message, state)


@dp.message_handler(lambda message: message.text == "Следущий шаг")
async def preview_advertisement(message: types.Message):
    await car_bot.preview_advertisement(message)


@dp.message_handler(lambda message: message.text == "Отправить в канал")
async def send_advertisement(message: types.Message, state: FSMContext):
    await car_bot.send_advertisement(message)

@dp.message_handler(lambda message: message.text == "Отменить и заполнить заново")
async def fill_again(message: types.Message, state: FSMContext):
    await car_bot.fill_again(message, state)

@dp.message_handler(lambda message: message.text == "Добавить ещё объявление")
async def add_more(message: types.Message, state: FSMContext):
    await car_bot.add_more(message, state)

@dp.message_handler(lambda message: message.text == "Ускорить продажу")
async def promotion(message: types.Message, state: FSMContext):
    await car_bot.promotion(message, state)


# старт бота
if __name__ == '__main__':
    executor.start_polling(dp, skip_updates=True)
