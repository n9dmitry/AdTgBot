import asyncio
import json
import random
import datetime
import uuid
import openpyxl

from aiogram import Bot, Dispatcher, Router, F, types
from aiogram.types import KeyboardButton, InputMediaPhoto, Message
from aiogram.enums import ParseMode
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.fsm.context import FSMContext
from aiogram.client.session.aiohttp import AiohttpSession
from aiogram.filters import Command, CommandStart
from aiogram.utils.keyboard import ReplyKeyboardBuilder
from aiogram.utils.markdown import hlink
from fuzzywuzzy import fuzz

from config import *
from states import *
from validation import *
from enumlist import *
from middleware_photogroup import AlbumMiddleware

router = Router()

router.message.middleware(AlbumMiddleware())

lock = asyncio.Lock()
session = AiohttpSession()
bot_settings = {"session": session, "parse_mode": ParseMode.HTML}
bot = Bot(token=API_TOKEN)
storage = MemoryStorage()


async def main():
    dp = Dispatcher()
    dp.include_router(router)
    await dp.start_polling(bot)


# Загрузка JSON в начале скрипта
with open('dicts.json', 'r', encoding='utf-8') as file:
    dicts = json.load(file)

dict_start_brands = dicts.get("dict_start_brands", {})
dict_car_body_types = dicts.get("dict_car_body_types", {})
dict_car_engine_types = dicts.get("dict_car_engine_types", {})
dict_car_transmission_types = dicts.get("dict_car_transmission_types", {})
dict_car_colors = dicts.get("dict_car_colors", {})
dict_car_document_statuses = dicts.get("dict_car_document_statuses", {})
dict_car_owners = dicts.get("dict_car_owners", {})
dict_car_customs_cleared = dicts.get("dict_car_customs_cleared", {})
dict_currency = dicts.get("dict_currency", {})
dict_car_conditions = dicts.get("dict_car_conditions", {})
dict_car_mileages = dicts.get("dict_car_mileages", {})
dict_edit_buttons = dicts.get("dict_edit_buttons", {})


# Конец импорта json словарей
# Создание клавиатуры

def create_keyboard(button_texts):
    buttons = [KeyboardButton(text=text) for text in button_texts]
    builder = ReplyKeyboardBuilder()
    builder.add(*buttons).adjust(2)
    return builder

async def send_photo_with_caption(message, state, image_path, caption, builder=None):
    user_data = await state.get_data()
    reply_markup = None
    if builder:
        reply_markup = builder.as_markup(resize_keyboard=True)
    await message.answer_photo(photo=types.FSInputFile(image_path), caption=caption, reply_markup=reply_markup)


async def recognize_car_model(message, brand_name):
    models = []
    similar_brands = []
    if brand_name.lower() in ['жигули']:
        brand_name = 'Lada (ВАЗ)'

    with open('cars.json', encoding='utf-8') as file:
        data = json.load(file)

    found_brand = False
    for item in data:
        # Сравнение имени бренда с учетом расстояния Левенштейна
        if 'name' in item and fuzz.token_sort_ratio(brand_name.lower(), item['name'].lower()) >= 90:
            if 'models' in item:
                models = item['models']
            found_brand = True
            break
        # Сравнение кириллического имени бренда
        elif 'cyrillic-name' in item and fuzz.token_sort_ratio(brand_name.lower(), item['cyrillic-name'].lower()) >= 90:
            if 'models' in item:
                models = item['models']
            found_brand = True
            break

    if not found_brand and len(brand_name) >= 3:
        for inner_item in data:
            # Поиск похожих брендов с учетом расстояния Левенштейна
            if 'name' in inner_item and fuzz.token_sort_ratio(brand_name.lower(), inner_item['name'].lower()) >= 75 and \
                    inner_item['name'] not in similar_brands:
                similar_brands.append(inner_item['name'])
            # Поиск похожих кириллических брендов
            elif 'cyrillic-name' in inner_item and fuzz.token_sort_ratio(brand_name.lower(),
                                                                         inner_item['cyrillic-name'].lower()) >= 75 \
                    and inner_item['name'] not in similar_brands:
                similar_brands.append(inner_item['name'])

        if similar_brands:
            response_message = "Похожие бренды:\n" + "\n".join(similar_brands)
            await message.answer(response_message)

    return models


# Команды
@router.message(F.text == "Перезагрузить бота")
@router.message(F.text == "Добавить ещё объявление")
@router.message(F.text == "Отменить и заполнить заново")
@router.message(User.STATE_SUPPORT_END)
@router.message(Command("restart"))
async def restart(message: types.Message, state: FSMContext):
    await state.clear()
    await message.answer("Бот перезапущен.")
    await start(message, state)


@router.message(Command("support"))
async def support(message: types.Message, state: FSMContext):
    user_data = await state.get_data()
    secret_number = str(random.randint(100, 999))

    await message.answer(f"Нашли баг? Давайте отправим сообщение разработчикам! "
                         f"Но перед этим введите проверку. Докажите что вы не робот. Напишите число {secret_number}:")
    user_data['secret_number'] = secret_number
    await state.update_data(user_data)
    await state.set_state(User.STATE_SUPPORT_VALIDATION)


@router.message(User.STATE_SUPPORT_VALIDATION)
async def support_validation(message: types.Message, state: FSMContext):
    user_data = await state.get_data()
    secret_number = user_data['secret_number']
    if message.text.isdigit() and message.text == secret_number:
        await message.reply(f"Проверка пройдена успешно!")
        await asyncio.sleep(1)
        await message.answer(f"Опишите техническую проблему в деталях для разработчиков: ")
        await state.set_state(User.STATE_SUPPORT_MESSAGE)
    else:
        await message.answer(f"Попробуйте ещё раз!")
        await asyncio.sleep(1)
        await support(message, state)


@router.message(User.STATE_SUPPORT_MESSAGE)
async def support_message(message: types.Message, state: FSMContext):
    current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    message_to_write = f"""
    Дата: {current_time}
    Имя: {message.from_user.full_name}
    Telegram @{message.from_user.username or message.from_user.id} 

    Сообщение: {message.text}
    ...
        """
    # Открываем файл для записи и записываем сообщение
    with open("support.txt", "a") as file:
        file.write(message_to_write)
    builder = create_keyboard(['Перезагрузить бота'])
    await message.reply("Спасибо за ваше сообщение! Мы рассмотрим вашу проблему!",
                        reply_markup=builder.as_markup(resize_keyboard=True))
    await state.set_state(User.STATE_SUPPORT_END)


# Начало работы бота
@router.message(CommandStart())
async def start(message: types.Message, state: FSMContext):
    image_hello_path = ImageDirectory.auto_say_hi
    await send_photo_with_caption(message, state, image_hello_path,
                                  f"Привет, {message.from_user.first_name}! Давай продадим твоё авто! Начнём же сбор данных!")
    await asyncio.sleep(0.5)
    builder = create_keyboard(dict_start_brands)
    image_path = ImageDirectory.auto_car_brand
    await send_photo_with_caption(message, state, image_path, "Выберите бренд автомобиля:", builder)
    await state.set_state(User.STATE_CAR_BRAND)


@router.message(User.STATE_CAR_BRAND)
async def get_car_brand(message, state):
    user_data = await state.get_data()
    search_brand = message.text
    await state.update_data(car_brand=search_brand)  # Обновляем данные пользователя в состоянии
    if search_brand == "⌨ Введите свой бренд":
        await message.answer("Пожалуйста, введите название марки своего автомобиля:")
    else:
        models = await recognize_car_model(message, search_brand)

        if not models:
            await message.answer("Марка не найдена, попробуйте еще раз")
            await state.set_state(User.STATE_CAR_BRAND)
        else:
            model_names = [model['name'] for model in models]
            builder = create_keyboard(model_names)
            await message.answer("Выберите модель автомобиля из списка:")
            await message.answer(f"Модели автомобилей марки '{search_brand}':",
                                 reply_markup=builder.as_markup(resize_keyboard=True))

            await state.set_state(User.STATE_CAR_MODEL)


@router.message(User.STATE_CAR_MODEL)
async def get_car_model(message, state):
    user_data = await state.get_data()
    print('0', user_data)
    await state.update_data(car_model=message.text)  # Обновляем данные пользователя в состоянии
    image_path = ImageDirectory.auto_car_year
    await send_photo_with_caption(message, state, image_path, "Какой год выпуска у автомобиля? (напишите)")
    await state.set_state(User.STATE_CAR_YEAR)  # Переключаемся на следующий шаг


@router.message(User.STATE_CAR_YEAR)
async def get_car_year(message, state):
    user_data = await state.get_data()
    print('1', user_data)
    if await validate_year(message.text):
        await state.update_data(car_year=message.text)
        builder = create_keyboard(dict_car_body_types)
        image_path = ImageDirectory.auto_car_body_type
        await send_photo_with_caption(message, state, image_path, "Отлично! Какой тип кузова у автомобиля?", builder)
        await state.set_state(User.STATE_CAR_BODY_TYPE)
    else:
        await message.answer("Пожалуйста, введите год в формате YYYY (например, 1990 или 2024)")
        await state.set_state(User.STATE_CAR_YEAR)


@router.message(User.STATE_CAR_BODY_TYPE)
async def get_car_body_type(message, state):
    user_data = await state.get_data()
    print('2', user_data)
    if await validate_button_input(message.text, dict_car_body_types):
        builder = create_keyboard(dict_car_engine_types)
        await state.update_data(car_body_type=message.text)
        image_path = ImageDirectory.auto_car_engine_type
        await send_photo_with_caption(message, state, image_path, "Отлично! Какой тип двигателя у автомобиля?", builder)
        await state.set_state(User.STATE_CAR_ENGINE_TYPE)
    else:
        builder = create_keyboard(dict_car_body_types)
        await message.answer("Пожалуйста, выберите корректный тип кузова.", builder.as_markup(resize_keyboard=True))
        await state.set_state(User.STATE_CAR_BODY_TYPE)


@router.message(User.STATE_CAR_ENGINE_TYPE)
async def get_car_engine_type(message, state):
    user_data = await state.get_data()
    print('3', user_data)
    if await validate_button_input(message.text, dict_car_engine_types):
        await state.update_data(car_engine_type=message.text)
        image_path = ImageDirectory.auto_car_engine_volume
        await send_photo_with_caption(message, state, image_path,
                                      "Хорошо! Какой объем двигателя у автомобиля (л.)? (напишите через точку: например 1.6)")
        await state.set_state(User.STATE_CAR_ENGINE_VOLUME)
    else:
        builder = create_keyboard(dict_car_engine_types)
        await message.answer("Пожалуйста, выберите корректный тип двигателя.",
                             reply_markup=builder.as_markup(resize_keyboard=True))
        await state.set_state(User.STATE_CAR_ENGINE_TYPE)


@router.message(User.STATE_CAR_ENGINE_VOLUME)
async def get_car_engine_volume(message, state):
    user_data = await state.get_data()
    print('4', user_data)
    try:
        if "," in message.text:  # Проверяем наличие запятой
            message.text = message.text.replace(',', '.')  # Заменяем запятую на точку
        volume = float(message.text)  # Преобразуем текст в число

        if await validate_engine_volume(volume) and 0.2 <= volume <= 10.0:
            await state.update_data(car_engine_volume=volume)
            image_path = ImageDirectory.auto_car_power
            await send_photo_with_caption(message, state, image_path,
                                          "Отлично! Укажите мощность двигателя автомобиля от 50 до 1000 (л.с.). (напишите)")
            await state.set_state(User.STATE_CAR_POWER)
        else:
            await message.answer(
                "Пожалуйста, корректный объем двигателя (в пределах от 0.2 до 10.0 литров) через точку или целым числом(!).")
            await state.set_state(User.STATE_CAR_ENGINE_VOLUME)

    except ValueError:
        # Если не удалось преобразовать введенный текст в число
        await message.answer(
            "Пожалуйста, корректный объем двигателя (в пределах от 0.2 до 10.0 литров) через точку или целым числом(!).")
        await state.set_state(User.STATE_CAR_ENGINE_VOLUME)


@router.message(User.STATE_CAR_POWER)
async def get_car_power(message, state):
    user_data = await state.get_data()
    print('5', user_data)
    if await validate_car_power(message.text):
        builder = create_keyboard(dict_car_transmission_types)
        await state.update_data(car_power=message.text)
        image_path = ImageDirectory.auto_car_transmission_type
        await send_photo_with_caption(message, state, image_path,
                                      "Отлично! Какой тип коробки передач используется в автомобиле?", builder)
        await state.set_state(User.STATE_CAR_TRANSMISSION_TYPE)
    else:
        await message.answer(
            "Пожалуйста, введите корректную мощность двигателя (в пределах от 50 до 1000 л.с.).")
        await state.set_state(User.STATE_CAR_POWER)


@router.message(User.STATE_CAR_TRANSMISSION_TYPE)
async def get_car_transmission_type(message, state):
    user_data = await state.get_data()
    print('6', user_data)
    if await validate_button_input(message.text, dict_car_transmission_types):
        builder = create_keyboard(dict_car_colors)
        await state.update_data(car_transmission_type=message.text)
        image_path = ImageDirectory.auto_car_color
        await send_photo_with_caption(message, state, image_path, "Какого цвета автомобиль?", builder)
        await state.set_state(User.STATE_CAR_COLOR)
    else:
        builder = create_keyboard(dict_car_transmission_types)
        await message.answer("Пожалуйста, выберите корректный тип трансмиссии.",
                             reply_markup=builder.as_markup(resize_keyboard=True))
        await state.set_state(User.STATE_CAR_TRANSMISSION_TYPE)


@router.message(User.STATE_CAR_COLOR)
async def get_car_color(message, state):
    user_data = await state.get_data()
    print('7', user_data)
    if await validate_button_input(message.text, dict_car_colors):
        builder = create_keyboard(dict_car_mileages)
        await state.update_data(car_color=message.text)
        image_path = ImageDirectory.auto_car_mileage
        await send_photo_with_caption(message, state, image_path,
                                      "Каков пробег автомобиля(км.)? (если новый, выберите 'Новый')", builder)
        await state.set_state(User.STATE_CAR_MILEAGE)
    else:
        builder = create_keyboard(dict_car_colors)
        await message.answer("Пожалуйста, выберите корректный цвет автомобиля.",
                             reply_markup=builder.as_markup(resize_keyboard=True))
        await state.set_state(User.STATE_CAR_COLOR)


@router.message(User.STATE_CAR_MILEAGE)
async def get_car_mileage(message, state):
    user_data = await state.get_data()
    print('8', user_data)
    if await validate_car_mileage(message.text):
        builder = create_keyboard(dict_car_document_statuses)
        await state.update_data(car_mileage=message.text)
        image_path = ImageDirectory.auto_car_document_status
        await send_photo_with_caption(message, state, image_path, "Каков статус документов у автомобиля ?", builder)
        await state.set_state(User.STATE_CAR_DOCUMENT_STATUS)
    else:
        builder = create_keyboard(dict_car_mileages)
        await message.answer("Пожалуйста, введите корректное значение пробега.",
                             reply_markup=builder.as_markup(resize_keyboard=True))
        await state.set_state(User.STATE_CAR_MILEAGE)


@router.message(User.STATE_CAR_DOCUMENT_STATUS)
async def get_car_document_status(message, state):
    user_data = await state.get_data()
    print('9', user_data)
    if await validate_button_input(message.text, dict_car_document_statuses):

        builder = create_keyboard(dict_car_owners)
        await state.update_data(car_document_status=message.text)
        image_path = ImageDirectory.auto_car_owners
        await send_photo_with_caption(message, state, image_path, "Сколько владельцев у автомобиля?", builder)
        await state.set_state(User.STATE_CAR_OWNERS)
    else:
        builder = create_keyboard(dict_car_document_statuses)
        await message.answer("Пожалуйста, выберите корректный статус документов автомобиля.",
                             reply_markup=builder.as_markup(resize_keyboard=True))
        await state.set_state(User.STATE_CAR_DOCUMENT_STATUS)


@router.message(User.STATE_CAR_OWNERS)
async def get_car_owners(message, state):
    user_data = await state.get_data()
    print('10', user_data)
    if await validate_button_input(message.text, dict_car_owners):
        builder = create_keyboard(dict_car_customs_cleared)
        await state.update_data(car_owners=message.text)
        image_path = ImageDirectory.auto_car_customs_cleared
        await send_photo_with_caption(message, state, image_path, "Растаможен ли автомобиль?", builder)
        await state.set_state(User.STATE_CAR_CUSTOMS_CLEARED)
    else:
        builder = create_keyboard(dict_car_owners)
        await message.answer("Пожалуйста, выберите корректное количество владельцев автомобиля.",
                             reply_markup=builder.as_markup(resize_keyboard=True))
        await state.set_state(User.STATE_CAR_OWNERS)


@router.message(User.STATE_CAR_CUSTOMS_CLEARED)
async def get_car_customs_cleared(message, state):
    user_data = await state.get_data()
    print('11', user_data)
    if await validate_button_input(message.text, dict_car_customs_cleared):
        builder = create_keyboard(dict_car_conditions)
        await state.update_data(car_customs_cleared=message.text)
        image_path = ImageDirectory.auto_car_condition
        await send_photo_with_caption(message, state, image_path, "Выберите состояние автомобиля:", builder)
        await state.set_state(User.STATE_CAR_CONDITION)
    else:
        builder = create_keyboard(dict_car_customs_cleared)
        await message.answer("Пожалуйста, выберите корректный статус растаможки автомобиля.",
                             reply_markup=builder.as_markup(resize_keyboard=True))
        await state.set_state(User.STATE_CAR_CUSTOMS_CLEARED)


@router.message(User.STATE_CAR_CONDITION)
async def get_car_condition(message, state):
    user_data = await state.get_data()
    print('12', user_data)
    if await validate_button_input(message.text, dict_car_conditions):
        await state.update_data(car_condition=message.text)
        image_path = ImageDirectory.auto_car_description
        await send_photo_with_caption(message, state, image_path, "Описание автомобиля. (напишите до 350 символов)")
        await state.set_state(User.STATE_CAR_DESCRIPTION)
    else:
        builder = create_keyboard(dict_car_conditions)
        await message.answer("Пожалуйста, выберите корректное состояние автомобиля.",
                             reply_markup=builder.as_markup(resize_keyboard=True))
        await state.set_state(User.STATE_CAR_CONDITION)


@router.message(User.STATE_CAR_DESCRIPTION)
async def get_car_description(message, state):
    user_data = await state.get_data()
    print('13', user_data)
    if await validate_length_text(message):
        if await validate_car_description(message.text):
            builder = create_keyboard(dict_currency)
            await state.update_data(car_description=message.text)
            image_path = ImageDirectory.auto_car_currency
            await send_photo_with_caption(message, state, image_path, "Выберите валюту:", builder)
            await state.set_state(User.STATE_SELECT_CURRENCY)
        else:
            await message.answer("Пожалуйста, введите корректное описание.")
            await state.set_state(User.STATE_CAR_DESCRIPTION)
    else:
        await message.answer("Ваше описание сильно большое. Напишите до ~350 символов:")
        await state.set_state(User.STATE_CAR_DESCRIPTION)


@router.message(User.STATE_SELECT_CURRENCY)
async def select_currency(message, state):
    user_data = await state.get_data()
    print('14', user_data)
    if await validate_button_input(message.text, dict_currency):
        await state.update_data(currency=message.text)
        image_path = ImageDirectory.auto_car_price
        await send_photo_with_caption(message, state, image_path, "Цена автомобиля?")
        await state.set_state(User.STATE_CAR_PRICE)
    else:
        builder = create_keyboard(dict_currency)
        await message.answer("Пожалуйста, выберите корректную валюту.",
                             reply_markup=builder.as_markup(resize_keyboard=True))
        await state.set_state(User.STATE_SELECT_CURRENCY)


@router.message(User.STATE_CAR_PRICE)
async def get_car_price(message, state):
    user_data = await state.get_data()
    print('15', user_data)
    if await validate_car_price(message.text):
        await state.update_data(car_price=message.text)
        image_path = ImageDirectory.auto_car_location
        await send_photo_with_caption(message, state, image_path,
                                      "Прекрасно! Где находится автомобиль? Город/пункт. (напишите)")
        await state.set_state(User.STATE_CAR_LOCATION)
    else:
        await message.answer("Пожалуйста, введите корректную цену.")
        await state.set_state(User.STATE_CAR_PRICE)


@router.message(User.STATE_CAR_LOCATION)
async def get_car_location(message, state):
    user_data = await state.get_data()
    print('16', user_data)
    if await validate_car_location(message.text):
        await state.update_data(car_location=message.text)
        image_path = ImageDirectory.auto_seller_name
        await send_photo_with_caption(message, state, image_path, "Прекрасно! Укажите имя продавца. (напишите)")
        await state.set_state(User.STATE_SELLER_NAME)
    else:
        await message.answer("Пожалуйста, введите корректные данные.")
        await state.set_state(User.STATE_CAR_LOCATION)


@router.message(User.STATE_SELLER_NAME)
async def get_seller_name(message, state):
    user_data = await state.get_data()
    print('17', user_data)
    if await validate_name(message.text) is True:
        await state.update_data(seller_name=message.text)
        image_path = ImageDirectory.auto_seller_phone
        await send_photo_with_caption(message, state, image_path,
                                      "Отлично! Какой телефонный номер у продавца? (напишите в формате +7XXXNNNXXNN или 8XXXNNNXXNN)")
        await state.set_state(User.STATE_SELLER_PHONE)
    else:
        await message.answer("Пожалуйста, введите корректное имя.")
        await state.set_state(User.STATE_SELLER_NAME)


@router.message(User.STATE_SELLER_PHONE)
async def get_seller_phone(message, state):
    user_data = await state.get_data()
    print('18', user_data)
    if await validate_phone_number(message.text) is True:
        phone_text = '+7' + message.text[1:] if message.text.startswith('8') else message.text
        await state.update_data(seller_phone=phone_text)
        if await validate_final_length(message, state, user_data):
            image_path = ImageDirectory.auto_car_photos
            await send_photo_with_caption(message, state, image_path,
                                          "Добавьте фотографии авто до 10 штук (За один раз!)")
            await state.set_state(User.STATE_CAR_PHOTO)
        else:
            await message.reply(
                f"Ваше сообщение получилось сильно большим! \nПерезагрузите бота и напишите объявление заново.")

    else:
        await message.answer("Пожалуйста, введите корректный номер в формате +7XXXNNNXXNN.")
        await state.set_state(User.STATE_SELLER_PHONE)


@router.message(User.STATE_CAR_PHOTO)
@router.message(F.media_group_id)
async def handle_photos(message: types.Message, state: FSMContext, album: list[Message]):
    user_data = await state.get_data()
    print('19', user_data)
    if 'sent_photos' not in user_data:
        user_data['sent_photos'] = []
    new_id = str(uuid.uuid4().int)[:6]
    if 'new_id' not in user_data:
        user_data['new_id'] = new_id

    print(user_data)
    caption = (
        f"🛞 <b>#{user_data['car_brand']}-{user_data['car_model']}</b>\n\n"
        f"   <b>-Год:</b> {user_data['car_year']}\n"
        f"   <b>-Пробег (км.):</b> {user_data['car_mileage']}\n"
        f"   <b>-Тип КПП:</b> {user_data['car_transmission_type']}\n"
        f"   <b>-Кузов:</b> {user_data['car_body_type']}\n"
        f"   <b>-Тип двигателя:</b> {user_data['car_engine_type']}\n"
        f"   <b>-Объем двигателя (л.):</b> {user_data['car_engine_volume']}\n"
        f"   <b>-Мощность (л.с.):</b> {user_data['car_power']}\n"
        f"   <b>-Цвет:</b> {user_data['car_color']}\n"
        f"   <b>-Статус документов:</b> {user_data['car_document_status']}\n"
        f"   <b>-Количество владельцев:</b> {user_data['car_owners']}\n"
        f"   <b>-Растаможка:</b> {'Да' if user_data['car_customs_cleared'] else 'Нет'}\n"
        f"   <b>-Состояние:</b> {user_data['car_condition']}\n\n"
        f"ℹ️<b>Дополнительная информация:</b> {user_data['car_description']}\n\n"
        f"🔥<b>Цена:</b> {user_data['car_price']} {user_data['currency']}\n\n"
        f"📍<b>Местоположение:</b> {user_data['car_location']}\n"
        f"👤<b>Продавец:</b> <span class='tg-spoiler'> {user_data['seller_name']} </span>\n"
        f"📲<b>Телефон продавца:</b> <span class='tg-spoiler'>{user_data['seller_phone']} </span>\n"
        f"💬<b>Телеграм:</b> <span class='tg-spoiler'>@{message.from_user.username if message.from_user.username is not None else 'по номеру телефона'}</span>\n\n"
        f" {hlink('Selbie Auto. Рынок тачек в ДНР', 'https://t.me/selbieauto')} | {hlink('Разместить авто', 'https://t.me/selbie_bot')} \n\n"
        f"<b>ID объявления: #{user_data['new_id']}</b>"

    )

    for message in album:
        if message.photo:
            top_photo = message.photo[-1]
            user_data['sent_photos'].append(
                InputMediaPhoto(media=top_photo.file_id, caption=None, parse_mode="HTML"))

    user_data['sent_photos'][0].caption = caption

    await state.update_data(user_data)

    builder = ReplyKeyboardBuilder([[types.KeyboardButton(text="Следущий шаг"), ]])
    if album:
        count_photos = len(album)
        await message.reply(f'{count_photos} Фото добавлены', reply_markup=builder.as_markup(resize_keyboard=True))

    await state.set_state(User.STATE_PREVIEW_ADVERTISMENT)


@router.message(F.text == "Отправить в канал")
async def send_advertisement(message: types.Message, state):
    user_data = await state.get_data()
    print('21', user_data)
    await add_data_to_excel(message, state)
    user_id = message.from_user.id
    await bot.send_media_group(chat_id=CHANNEL_ID, media=user_data['sent_photos'], disable_notification=True)
    builder = create_keyboard(['Добавить ещё объявление', 'Ускорить продажу'])
    await bot.send_message(user_id, "Объявление отправлено в канал!",
                           reply_markup=builder.as_markup(resize_keyboard=True))
    await state.clear()


# @router.message(F.text == "Отменить и заполнить заново")
# async def fill_again(message: types.Message, state: FSMContext):
#     user_data = await state.get_data()
#     builder = create_keyboard(list(dict_car_brands_and_models.keys()))
#     image_path = ImageDirectory.auto_car_brand
#     with open(image_path, "rb"):
#         await message.answer_photo(photo=types.FSInputFile(image_path), caption="Выберите бренд автомобиля:",
#                                    reply_markup=builder.as_markup(resize_keyboard=True))
#     user_data['sent_photos'].clear()
#     await state.clear()
#     await state.set_state(User.STATE_CAR_BRAND)


@router.message(F.text == "Ускорить продажу")
async def promotion(message: types.Message):
    builder = create_keyboard(['Перезагрузить бота'])
    await message.reply("Чтобы купить закреп напишите @selbie_adv",
                        reply_markup=builder.as_markup(resize_keyboard=True))


@router.message(User.STATE_PREVIEW_ADVERTISMENT)
async def preview_advertisement(message: types.Message, state: FSMContext):
    user_data = await state.get_data()
    print('20', user_data)
    await bot.send_media_group(chat_id=message.chat.id, media=user_data['sent_photos'])

    builder = ReplyKeyboardBuilder([[
        KeyboardButton(text="Отправить в канал"),
        KeyboardButton(text="Отменить и заполнить заново")
    ]])

    await message.reply(
        "Так будет выглядеть ваше объявление. Вы можете либо разместить либо отменить и заполнить заново.",
        reply_markup=builder.as_markup(resize_keyboard=True))

    # db_fix.clear()


async def add_data_to_excel(message, state):
    user_data = await state.get_data()
    print('22 excel', user_data)
    file_path = 'db.xlsx'
    row_data = [
        user_data['new_id'], datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        user_data['car_brand'], user_data['car_model'], user_data['car_year'], user_data['car_body_type'],
        user_data['car_engine_type'], user_data['car_engine_volume'], user_data['car_power'],
        user_data['car_transmission_type'], user_data['car_color'], user_data['car_mileage'],
        user_data['car_document_status'], user_data['car_owners'], user_data['car_customs_cleared'],
        user_data['car_condition'], user_data['car_description'], user_data['currency'], user_data['car_price'],
        user_data['car_location'], user_data['seller_name'], user_data['seller_phone'],
        message.from_user.username if message.from_user.username is not None else 'по номеру телефона',
    ]

    # Проверяем, существует ли файл Excel
    if os.path.exists(file_path):
        workbook = openpyxl.load_workbook(file_path)
    else:
        workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Проверяем, нужно ли добавить заголовки
    if sheet.max_row == 1:
        headers = [
            'ID', 'Дата', 'Бренд', 'Модель', 'Год', 'Тип кузова',
            'Тип двигателя', 'Объем двигателя (л)', 'Мощность (л.с.)', 'Тип трансмиссии',
            'Цвет', 'Пробег (км)', 'Статус документа', 'Количество владельцев', 'Растаможен',
            'Состояние', 'Дополнительное описание', 'Валюта', 'Цена',
            'Местоположение', 'Имя продавца', 'Телефон продавца', 'Телеграм'

        ]
        sheet.append(headers)

    sheet.append(row_data)
    workbook.save(file_path)


# end support


# старт бота
if __name__ == '__main__':
    asyncio.run(main())
