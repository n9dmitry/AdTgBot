# pip install aiogram==3.4.1

import asyncio
from aiogram import Bot, Dispatcher, Router, F, types
from aiogram.types import KeyboardButton, InputMediaPhoto
from aiogram.enums import ParseMode
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.fsm.context import FSMContext
from aiogram.client.session.aiohttp import AiohttpSession
from aiogram.filters import Command, CommandStart
from aiogram.utils.keyboard import ReplyKeyboardBuilder
import random
import datetime
import uuid
import openpyxl
from config import *
from states import *
from validation import *
import json
from enumlist import *

router = Router(name=__name__)

lock = asyncio.Lock()

session = AiohttpSession()
bot_settings = {"session": session, "parse_mode": ParseMode.HTML}
bot = Bot(token=API_TOKEN)
storage=MemoryStorage()

buffered_photos = []

async def main():
    dp = Dispatcher()
    dp.include_router(router)
    await dp.start_polling(bot)

# Загрузка JSON в начале скрипта
with open('dicts.json', 'r', encoding='utf-8') as file:
    dicts = json.load(file)

dict_car_brands_and_models = dicts.get("dict_car_brands_and_models", {})
dict_car_body_types = dicts.get("dict_car_body_types", {})
dict_car_engine_types = dicts.get("dict_car_engine_types", {})
dict_car_transmission_types = dicts.get("dict_car_transmission_types", {})
dict_car_colors = dicts.get("dict_car_colors", {})
dict_car_document_statuses = dicts.get("dict_car_document_statuses", {})
dict_car_owners = dicts.get("dict_car_owners", {})
dict_car_customs_cleared = dicts.get("dict_car_customs_cleared", {})
dict_currency = dicts.get("dict_currency", {})
dict_car_conditions = dicts.get("dict_car_conditions", {})
dict_car_mileages = dicts.get("dict_car_mileages", {})
dict_edit_buttons = dicts.get("dict_edit_buttons", {})
# Конец импорта json словарей


# Создание клавиатуры
def create_keyboard(button_texts):
    buttons = [KeyboardButton(text=text) for text in button_texts]
    builder = ReplyKeyboardBuilder()
    builder.add(*buttons)
    return builder


# class CarBotHandler:
def __init__(self):
    self.lock = asyncio.Lock()
    # self.car_bot = CarBotHandler()

# Команды
@router.message(F.text == "Перезагрузить бота")
async def restart(message: types.Message, state: FSMContext):
    await state.clear()
    await message.answer("Бот перезапущен.")
    await bot.start(message, state)

@router.message(Command("support"))
async def support(message: types.Message, state: FSMContext):
    await state.clear()
    secret_number = str(random.randint(100, 999))

    await message.answer(f"Нашли баг? Давайте отправим сообщение разработчикам! "
                         f"Но перед этим введите проверку. Докажите что вы не робот. Напишите число {secret_number}:")
    await state.set_state(User.STATE_SUPPORT_VALIDATION)

@router.message(User.STATE_SUPPORT_VALIDATION)
async def support_validation(message: types.Message, state: FSMContext, secret_number):
    if message.text.isdigit() and message.text == secret_number:
        await message.reply(f"Проверка пройдена успешно!")
        await asyncio.sleep(1)
        await message.answer(f"Опишите техническую проблему в деталях для разработчиков: ")
        await state.set_state(User.STATE_SUPPORT_MESSAGE)
    else:
        await message.answer(f"Попробуйте ещё раз!")
        await asyncio.sleep(1)
        await bot.support(message, state)

@router.message(User.STATE_SUPPORT_MESSAGE)
async def support_message(self, message: types.Message, state: FSMContext):
    current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    message_to_write = f"""
    Дата: {current_time}
    Имя: {message.from_user.full_name}
    Telegram @{message.from_user.username or message.from_user.id} 
  
    Сообщение: {message.text}
    ...
        """

    # Открываем файл для записи и записываем сообщение
    with open("support.txt", "a") as file:
        file.write(message_to_write)
    builder = create_keyboard(['Перезагрузить бота'])
    await message.reply("Спасибо за ваше сообщение! Мы рассмотрим вашу проблему!", reply_markup=builder.as_markup(resize_keyboard=True))
    await state.set_state(User.STATE_SUPPORT_END)

@router.message(User.STATE_SUPPORT_END)
async def support_end(self, message: types.Message, state: FSMContext):
    await self.car_bot.restart(message, state)


# Начало работы бота

@router.message(CommandStart())
async def start(self, message: types.Message, state: FSMContext):
    image_hello_path = ImageDirectory.auto_say_hi
    await message.answer_photo(photo=types.FSInputFile(image_hello_path), caption=f"Привет, {message.from_user.first_name}! Давай продадим твоё авто! Начнём же сбор данных!")
    await asyncio.sleep(0.1)
    builder = create_keyboard(list(dict_car_brands_and_models.keys()))
    image_path = ImageDirectory.auto_car_brand
    await message.answer_photo(photo=types.FSInputFile(image_path), caption="Выберите бренд автомобиля:", reply_markup=builder.as_markup(resize_keyboard=True, row_width=2))
    await state.set_state(User.STATE_CAR_BRAND)

@router.message(User.STATE_CAR_BRAND)
async def get_car_brand(message: types.Message, state: FSMContext):
    user_data = (await state.get_data()).get("user_data", {})
    selected_brand = message.text
    valid_brands = dict_car_brands_and_models
    if await validate_car_brand(selected_brand, valid_brands):
        user_data["car_brand"] = selected_brand
        await state.update_data(user_data=user_data)
        builder = create_keyboard(
            dict_car_brands_and_models[selected_brand])
        image_path = ImageDirectory.auto_car_model
        await message.answer_photo(photo=types.FSInputFile(image_path), caption="Отлично! Выберите модель автомобиля:",
                                   reply_markup=builder.as_markup(resize_keyboard=True, row_width=2))
        # with open(image_path, "rb"):
        #     await message.send_photo(message.from_user.id, image_path, caption="Отлично! Выберите модель автомобиля:", reply_markup=builder.as_markup(resize_keyboard=True))
        await state.set_state(User.STATE_CAR_PHOTO)
    else:
        builder = create_keyboard(dict_car_brands_and_models.keys())
        await bot.send_message(message.from_user.id, "Пожалуйста, выберите бренд из предложенных вариантов или напишите нам если вашего бренда нет", reply_markup=builder.as_markup(resize_keyboard=True))
        await state.set_state(User.STATE_CAR_PHOTO)

@router.message(User.STATE_CAR_PHOTO)
@router.message(F.photo)
async def handle_photos(self, message: types.Message, state: FSMContext):
    user_data = await state.get_data()
    photo_id = message.photo[-1].file_id


    self.new_id = str(uuid.uuid4().int)[:6]

    caption = (
        f"🛞 <b>#{user_data.get('user_data').get('car_brand')}-"
        f"💬<b>Телеграм:</b> <span class='tg-spoiler'>{message.from_user.username if message.from_user.username is not None else 'по номеру телефона'}</span>\n\n"
        f"ООО 'Продвижение' Авто в ДНР (link: разместить авто)\n\n"
        f"<b>ID объявления: #{self.new_id}</b>"
    )




    if "sent_photos" not in user_data:
        user_data["sent_photos"] = []

    user_data["sent_photos"].append(
        {"file_id": photo_id,})
    buffered_photos.append(InputMediaPhoto(
        media=photo_id, caption=caption, parse_mode=ParseMode.HTML))
    if len(buffered_photos) > 1:
        for i in range(len(buffered_photos) - 1):
            buffered_photos[i].caption = None
        last_photo = buffered_photos[-1]
        last_photo.caption = caption

    builder = ReplyKeyboardBuilder(
        [
            [
                types.KeyboardButton(text="Следущий шаг"),
            ]
        ]
    )

    await message.answer("Фото добавлено", reply_markup=builder.as_markup(resize_keyboard=True))

    self.db_fix = user_data
    await state.clear()

async def add_data_to_excel(self, message):
    file_path = 'db.xlsx'

    row_data = [
        self.new_id,
        datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        self.db_fix.get('user_data').get('car_brand', ''),
        self.db_fix.get('user_data').get('car_model', ''),
        self.db_fix.get('user_data').get('car_year', ''),
        self.db_fix.get('user_data').get('car_mileage', ''),
        self.db_fix.get('user_data').get('car_transmission_type', ''),
        self.db_fix.get('user_data').get('car_body_type', ''),
        self.db_fix.get('user_data').get('car_engine_type', ''),
        self.db_fix.get('user_data').get('car_engine_volume', ''),
        self.db_fix.get('user_data').get('car_power', ''),
        self.db_fix.get('user_data').get('car_color', ''),
        self.db_fix.get('user_data').get('car_document_status', ''),
        self.db_fix.get('user_data').get('car_owners', ''),
        self.db_fix.get('user_data').get('car_customs_cleared'),
        self.db_fix.get('user_data').get('car_condition', ''),
        self.db_fix.get('user_data').get('car_description', ''),
        self.db_fix.get('user_data').get('car_price', ''),
        self.db_fix.get('user_data').get('currency', ''),
        self.db_fix.get('user_data').get('car_location', ''),
        self.db_fix.get('user_data').get('seller_name', ''),
        self.db_fix.get('user_data').get('seller_phone', ''),
        message.from_user.username if message.from_user.username is not None else 'по номеру телефона',
    ]

    # Проверяем, существует ли файл Excel
    if os.path.exists(file_path):
        workbook = openpyxl.load_workbook(file_path)
    else:
        workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Проверяем, нужно ли добавить заголовки
    if sheet.max_row == 1:
        headers = [
            'ID','Дата', 'Бренд', 'Модель', 'Год', 'Пробег (км)', 'Тип трансмиссии',
            'Тип кузова', 'Тип двигателя', 'Объем двигателя (л)', 'Мощность (л.с.)',
            'Цвет', 'Статус документа', 'Количество владельцев', 'Растаможен',
            'Состояние', 'Дополнительное описание', 'Цена', 'Валюта',
            'Местоположение', 'Имя продавца', 'Телефон продавца', 'Телеграм'
        ]
        sheet.append(headers)

    sheet.append(row_data)
    workbook.save(file_path)

@router.message(F.text == "Следущий шаг")
async def preview_advertisement(self, message: types.Message):
    await bot.send_media_group(chat_id=message.chat.id, media=buffered_photos, disable_notification=True)
    builder = ReplyKeyboardBuilder(
    [
        [
        KeyboardButton(text="Отправить в канал"),
        KeyboardButton(text="Отменить и заполнить заново")
        ]
    ]
    )
    await message.reply("Так будет выглядеть ваше объявление. Вы можете либо разместить либо отменить и заполнить заново.", reply_markup=builder.as_markup(resize_keyboard=True))

@router.message(F.text == "Отправить в канал")
async def send_advertisement(self, message: types.Message):
    async with lock:
        user_id = message.from_user.id
        await message.add_data_to_excel(message)
        await bot.send_media_group(chat_id=CHANNEL_ID, media=buffered_photos, disable_notification=True)
        builder = create_keyboard(['Добавить ещё объявление', 'Ускорить продажу'])
        await bot.send_message(user_id, "Объявление отправлено в канал!", reply_markup=builder.as_markup(resize_keyboard=True))

        buffered_photos.clear()

@router.message(F.text == "Отменить и заполнить заново")
async def fill_again(message: types.Message, state: FSMContext):
    builder = create_keyboard(list(dict_car_brands_and_models.keys()))
    image_path = ImageDirectory.auto_car_brand # Путь к вашему изображению
    with open(image_path, "rb"):
        await message.answer_photo(image_path, caption="Выберите бренд автомобиля:", reply_markup=builder.as_markup(resize_keyboard=True))
    async with lock:
        buffered_photos.clear()
    await state.set_state(User.STATE_CAR_BRAND)

@router.message(F.text == "Добавить ещё объявление")
async def add_more(self, message: types.Message, state: FSMContext):
    await self.car_bot.restart(message, state)

@router.message(F.text == "Ускорить продажу")
async def promotion(self, message: types.Message):
    builder = create_keyboard(['Перезагрузить бота'])
    await message.reply("Чтобы купить закреп напишите @n9dmitry", reply_markup=builder.as_markup(resize_keyboard=True))
# end support


# старт бота
if __name__ == '__main__':
    asyncio.run(main())
