# from main import *
import asyncio
import random
import datetime
import uuid
import openpyxl
import json

from aiogram import Bot, Dispatcher, executor, types
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, InputMediaPhoto
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
from aiogram.utils.markdown import hlink

from config import *
from states import *
from enumlist import *


# Загрузка JSON в начале скрипта
with open('dicts.json', 'r', encoding='utf-8') as file:
    dicts = json.load(file)

dict_car_brands_and_models = dicts.get("dict_car_brands_and_models", {})
dict_car_body_types = dicts.get("dict_car_body_types", {})
dict_car_engine_types = dicts.get("dict_car_engine_types", {})
dict_car_transmission_types = dicts.get("dict_car_transmission_types", {})
dict_car_colors = dicts.get("dict_car_colors", {})
dict_car_document_statuses = dicts.get("dict_car_document_statuses", {})
dict_car_owners = dicts.get("dict_car_owners", {})
dict_car_customs_cleared = dicts.get("dict_car_customs_cleared", {})
dict_currency = dicts.get("dict_currency", {})
dict_car_conditions = dicts.get("dict_car_conditions", {})
dict_car_mileages = dicts.get("dict_car_mileages", {})
dict_edit_buttons = dicts.get("dict_edit_buttons", {})
# Конец импорта json словарей


# Создание клавиатуры
def create_keyboard(button_texts, resize_keyboard=True):
    keyboard = ReplyKeyboardMarkup(
        resize_keyboard=resize_keyboard, row_width=2)
    buttons = [KeyboardButton(text=text) for text in button_texts]
    keyboard.add(*buttons)
    return keyboard




class HRBotHandler:
    def __init__(self):
        self.lock = asyncio.Lock()


# Команды

    async def restart(self, event, state):
        await state.finish()
        await event.answer("Бот перезапущен.")
        await self.start(event, state)

    async def support(self, event, state):
        await state.finish()
        self.secret_number = str(random.randint(100, 999))

        await event.answer(f"Нашли баг? Давайте отправим сообщение разработчикам! "
                             f"Но перед этим введите проверку. Докажите что вы не робот. Напишите число {self.secret_number}:")
        await state.set_state(User.STATE_SUPPORT_VALIDATION)

    async def support_validation(self, event, state):
        if event.text.isdigit() and event.text == self.secret_number:
            await event.reply(f"Проверка пройдена успешно!")
            await asyncio.sleep(1)
            await event.answer(f"Опишите техническую проблему в деталях для разработчиков: ")
            await state.set_state(User.STATE_SUPPORT_MESSAGE)
        else:
            await event.answer(f"Попробуйте ещё раз!")
            await asyncio.sleep(1)
            await cmd_support(event, state)

    async def support_message(self, event: types.Message, state):
        # Получаем текущую дату и время
        current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        # Формируем строку для записи в файл
        message_to_write = f"""
        Дата: {current_time}
        Имя: {event.from_user.full_name}
        Telegram @{event.from_user.username or event.from_user.id} 
      
        Сообщение: {event.text}
        ...
            """

        # Открываем файл для записи и записываем сообщение
        with open("support.txt", "a") as file:
            file.write(message_to_write)
        keyboard = create_keyboard(['Перезагрузить бота'])
        await event.reply("Спасибо за ваше сообщение! Мы рассмотрим вашу проблему!", reply_markup=keyboard)
        await state.set_state(User.STATE_SUPPORT_END)
    async def support_end(selfself, event, state):
        if event.text == 'Перезагрузить бота':
            await cmd_restart(event, state)
        await state.finish()



# Начало работы бота

    async def start(self, event, state):
        image_hello_path = ImageDirectory.auto_say_hi
        with open(image_hello_path, "rb") as image_hello:
            self.m = await event.answer_photo(image_hello,
                                     caption=f"(найм) Привет, {event.from_user.first_name}! Давай продадим твоё авто! Начнём же сбор данных!")
        await asyncio.sleep(0)
        # self.m = await event.answer(f"Привет, {event.from_user.first_name}! Я бот для сбора данных. Давай начнем.")
        # keyboard = create_keyboard(list(dict_car_brands_and_models.keys()))
        # image_path = ImageDirectory.auto_car_brand  # Путь к вашему изображению
        # with open(image_path, "rb") as image:
        #     self.m = await event.answer_photo(image, caption="Выберите бренд автомобиля:", reply_markup=keyboard)
        # self.m = await event.answer("Выберите бренд автомобиля:", reply_markup=keyboard)
        await event.answer("Теперь фото:")

        await state.set_state(User.STATE_HR_PHOTO)


    async def handle_photos3(self, event, state):
        user_data = await state.get_data('user_data')
        photo_id = event.photo[-1].file_id


        self.new_id = str(uuid.uuid4().int)[:6]

        caption = (
            f"🛞 <b>#{user_data.get('user_data').get('car_brand')}-{user_data.get('user_data').get('car_model')}</b>\n\n"
            f"   <b>-Год:</b> {user_data.get('user_data', {}).get('car_year')}\n"
            f"   <b>-Пробег (км.):</b> {user_data.get('user_data').get('car_mileage')}\n"
            f"   <b>-Тип КПП:</b> {user_data.get('user_data').get('car_transmission_type')}\n"
            f"   <b>-Кузов:</b> {user_data.get('user_data').get('car_body_type')}\n"
            f"   <b>-Тип двигателя:</b> {user_data.get('user_data').get('car_engine_type')}\n"
            f"   <b>-Объем двигателя (л.):</b> {user_data.get('user_data').get('car_engine_volume')}\n"
            f"   <b>-Мощность (л.с.):</b> {user_data.get('user_data').get('car_power')}\n"
            f"   <b>-Цвет:</b> {user_data.get('user_data').get('car_color')}\n"
            f"   <b>-Статус документов:</b> {user_data.get('user_data').get('car_document_status')}\n"
            f"   <b>-Количество владельцев:</b> {user_data.get('user_data').get('car_owners')}\n"
            f"   <b>-Растаможка:</b> {'Да' if user_data.get('user_data').get('car_customs_cleared') else 'Нет'}\n"
            f"   <b>-Состояние:</b> {user_data.get('user_data').get('car_condition')}\n\n"
            f"ℹ️<b>Дополнительная информация:</b> {user_data.get('user_data').get('car_description')}\n\n"
            f"🔥<b>Цена:</b> {user_data.get('user_data').get('car_price')} {user_data.get('user_data').get('currency')}\n\n"
            f"📍<b>Местоположение:</b> {user_data.get('user_data').get('car_location')}\n"
            f"👤<b>Продавец:</b> <span class='tg-spoiler'> {user_data.get('user_data').get('seller_name')} </span>\n"
            f"📲<b>Телефон продавца:</b> <span class='tg-spoiler'>{user_data.get('user_data').get('seller_phone')} </span>\n"
            f"💬<b>Телеграм:</b> <span class='tg-spoiler'>@{event.from_user.username if event.from_user.username is not None else 'по номеру телефона'}</span>\n\n"
            f" {hlink('Selbie Auto. Рынок тачек в ДНР', 'https://t.me/selbieauto')} | {hlink('Разместить авто', 'https://t.me/selbie_bot')} \n\n"
            f"<b>ID объявления: #{self.new_id}</b>"
        )


        if "sent_photos" not in user_data:
            user_data["sent_photos"] = []

        user_data["sent_photos"].append(
            {"file_id": photo_id,})
        buffered_photos.append(InputMediaPhoto(
            media=photo_id, caption=caption, parse_mode=types.ParseMode.HTML))
        # # await self.m.delete()
        if len(buffered_photos) > 1:
            for i in range(len(buffered_photos) - 1):
                buffered_photos[i].caption = None
            last_photo = buffered_photos[-1]
            last_photo.caption = caption


        keyboard = ReplyKeyboardMarkup(resize_keyboard=True).add(
            KeyboardButton("Следущий шаг")
        )



        self.m = await event.answer("Фото добавлено", reply_markup=keyboard)


        self.db_fix = user_data

        await state.finish()


    async def add_data_to_excel(self, event):
        file_path = 'db3.xlsx'


        row_data = [
            self.new_id,
            datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            self.db_fix.get('user_data').get('car_brand', ''),
            self.db_fix.get('user_data').get('car_model', ''),
            self.db_fix.get('user_data').get('car_year', ''),
            self.db_fix.get('user_data').get('car_mileage', ''),
            self.db_fix.get('user_data').get('car_transmission_type', ''),
            self.db_fix.get('user_data').get('car_body_type', ''),
            self.db_fix.get('user_data').get('car_engine_type', ''),
            self.db_fix.get('user_data').get('car_engine_volume', ''),
            self.db_fix.get('user_data').get('car_power', ''),
            self.db_fix.get('user_data').get('car_color', ''),
            self.db_fix.get('user_data').get('car_document_status', ''),
            self.db_fix.get('user_data').get('car_owners', ''),
            self.db_fix.get('user_data').get('car_customs_cleared'),
            self.db_fix.get('user_data').get('car_condition', ''),
            self.db_fix.get('user_data').get('car_description', ''),
            self.db_fix.get('user_data').get('car_price', ''),
            self.db_fix.get('user_data').get('currency', ''),
            self.db_fix.get('user_data').get('car_location', ''),
            self.db_fix.get('user_data').get('seller_name', ''),
            self.db_fix.get('user_data').get('seller_phone', ''),
            event.from_user.username if event.from_user.username is not None else 'по номеру телефона',
        ]

        # Проверяем, существует ли файл Excel
        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(file_path)
        else:
            workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Проверяем, нужно ли добавить заголовки
        if sheet.max_row == 1:
            headers = [
                'ID','Дата', 'Бренд', 'Модель', 'Год', 'Пробег (км)', 'Тип трансмиссии',
                'Тип кузова', 'Тип двигателя', 'Объем двигателя (л)', 'Мощность (л.с.)',
                'Цвет', 'Статус документа', 'Количество владельцев', 'Растаможен',
                'Состояние', 'Дополнительное описание', 'Цена', 'Валюта',
                'Местоположение', 'Имя продавца', 'Телефон продавца', 'Телеграм'
            ]
            sheet.append(headers)

        sheet.append(row_data)
        workbook.save(file_path)

    async def preview_advertisement(self, event):
        await bot.send_media_group(chat_id=event.chat.id, media=buffered_photos, disable_notification=True)

        keyboard = ReplyKeyboardMarkup(resize_keyboard=True).add(
            KeyboardButton("Отправить в канал"),
            KeyboardButton("Отменить и заполнить заново"),
        )
        await event.reply("Так будет выглядеть ваше объявление. Вы можете либо разместить либо отменить и заполнить заново.", reply_markup=keyboard)

    async def send_advertisement(self, event):
        # user_id = event.from_user.id
        # await self.m.delete()
        async with lock:
            user_id = event.from_user.id
            await self.add_data_to_excel(event)
            await bot.send_media_group(chat_id=CHANNEL_ID, media=buffered_photos, disable_notification=True)
            keyboard = create_keyboard(['Добавить ещё объявление', 'Ускорить продажу'])
            await bot.send_message(user_id, "Объявление отправлено в канал!", reply_markup=keyboard)

            buffered_photos.clear()


    async def fill_again(self, event, state):
        keyboard = create_keyboard(list(dict_car_brands_and_models.keys()))
        image_path = ImageDirectory.auto_car_brand # Путь к вашему изображению
        with open(image_path, "rb") as image:
            self.m = await event.answer_photo(image, caption="Выберите бренд автомобиля:", reply_markup=keyboard)
        # self.m = await event.answer("Выберите бренд автомобиля:", reply_markup=keyboard)
        async with lock:
            buffered_photos.clear()
        await state.set_state(User.STATE_CAR_BRAND)

    async def add_more(self, event, state):
        await hr_bot.restart(event, state)

    async def promotion(self, event, state):
        keyboard = create_keyboard(['Перезагрузить бота'])
        await event.reply("Чтобы купить закреп напишите @selbie_adv", reply_markup=keyboard)


hr_bot = HRBotHandler()
bot = Bot(token=API_TOKEN)
dp = Dispatcher(bot, storage=MemoryStorage())
lock = asyncio.Lock()
buffered_photos = []


@dp.message_handler(lambda message: message.text == "Перезагрузить бота", state='*')
@dp.message_handler(commands=['restart'], state='*')
async def cmd_restart(event: types.Message, state: FSMContext):
    await hr_bot.restart(event, state)


@dp.message_handler(commands=["start"])
async def cmd_start(event: types.Message, state: FSMContext):
    await hr_bot.start(event, state)

#support
@dp.message_handler(commands=['support'], state='*')
async def cmd_support(event: types.Message, state: FSMContext):
    await hr_bot.support(event, state)

@dp.message_handler(state=User.STATE_SUPPORT_VALIDATION)
async def support_validation(event: types.Message, state: FSMContext):
    await hr_bot.support_validation(event, state)

@dp.message_handler(state=User.STATE_SUPPORT_MESSAGE)
async def support_message(event: types.Message, state: FSMContext):
    await hr_bot.support_message(event, state)

@dp.message_handler(state=User.STATE_SUPPORT_END)
async def support_end(event: types.Message, state: FSMContext):
    await hr_bot.restart(event, state)
# end support

@dp.message_handler(state=User.STATE_HR_PHOTO, content_types=['photo'])
async def handle_photos3(event: types.Message, state: FSMContext):
    await hr_bot.handle_photos3(event, state)


@dp.message_handler(lambda message: message.text == "Следущий шаг")
async def preview_advertisement(event: types.Message):
    await hr_bot.preview_advertisement(event)


@dp.message_handler(lambda message: message.text == "Отправить в канал")
async def send_advertisement(event: types.Message, state: FSMContext):
    await hr_bot.send_advertisement(event)

@dp.message_handler(lambda message: message.text == "Отменить и заполнить заново")
async def fill_again(event: types.Message, state: FSMContext):
    await hr_bot.fill_again(event, state)

@dp.message_handler(lambda message: message.text == "Добавить ещё объявление")
async def add_more(event: types.Message, state: FSMContext):
    await hr_bot.add_more(event, state)

@dp.message_handler(lambda message: message.text == "Ускорить продажу")
async def promotion(event: types.Message, state: FSMContext):
    await hr_bot.promotion(event, state)


# старт бота
if __name__ == '__main__':
    executor.start_polling(dp, skip_updates=True)
