import asyncio
from aiogram import Bot, Dispatcher, Router, F, types
from aiogram.types import KeyboardButton, InputMediaPhoto
from aiogram.enums import ParseMode
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.fsm.context import FSMContext
from aiogram.client.session.aiohttp import AiohttpSession
from aiogram.filters import Command, CommandStart
from aiogram.utils.keyboard import ReplyKeyboardBuilder
import random
import datetime
import uuid
import openpyxl
from config import *
from states import *
# from validation import *
import json
from enumlist import *
from middleware_photogroup import AlbumMiddleware
from aiogram.types import Message
from fuzzywuzzy import fuzz

router = Router()

router.message.middleware(AlbumMiddleware())

lock = asyncio.Lock()
session = AiohttpSession()
bot_settings = {"session": session, "parse_mode": ParseMode.HTML}
bot = Bot(token=API_TOKEN)
storage = MemoryStorage()


async def main():
    dp = Dispatcher()
    dp.include_router(router)
    await dp.start_polling(bot)


# Загрузка JSON в начале скрипта
with open('dicts.json', 'r', encoding='utf-8') as file:
    dicts = json.load(file)

dict_car_brands_and_models = dicts.get("dict_car_brands_and_models", {})
dict_car_body_types = dicts.get("dict_car_body_types", {})
dict_car_engine_types = dicts.get("dict_car_engine_types", {})
dict_car_transmission_types = dicts.get("dict_car_transmission_types", {})
dict_car_colors = dicts.get("dict_car_colors", {})
dict_car_document_statuses = dicts.get("dict_car_document_statuses", {})
dict_car_owners = dicts.get("dict_car_owners", {})
dict_car_customs_cleared = dicts.get("dict_car_customs_cleared", {})
dict_currency = dicts.get("dict_currency", {})
dict_car_conditions = dicts.get("dict_car_conditions", {})
dict_car_mileages = dicts.get("dict_car_mileages", {})
dict_edit_buttons = dicts.get("dict_edit_buttons", {})


# Конец импорта json словарей

# Создание клавиатуры
def create_keyboard(button_texts):
    buttons = [KeyboardButton(text=text) for text in button_texts]
    builder = ReplyKeyboardBuilder()
    builder.add(*buttons)
    return builder


async def recognize_car_model(event, brand_name):
    models = []
    similar_brands = []
    if brand_name.lower() in ['жигули']:
        brand_name = 'Lada (ВАЗ)'
    with open('cars.json', encoding='utf-8') as file:
        data = json.load(file)
    found_brand = False
    for item in data:
        # Сравнение имени бренда с учетом расстояния Левенштейна
        if 'name' in item and fuzz.token_sort_ratio(brand_name.lower(), item['name'].lower()) >= 90:
            if 'models' in item:
                models = item['models']
            found_brand = True
            break
        # Сравнение кириллического имени бренда
        elif 'cyrillic-name' in item and fuzz.token_sort_ratio(brand_name.lower(), item['cyrillic-name'].lower()) >= 90:
            if 'models' in item:
                models = item['models']
            found_brand = True
            break
    if not found_brand and len(brand_name) >= 3:
        for inner_item in data:
            # Поиск похожих брендов с учетом расстояния Левенштейна
            if 'name' in inner_item and fuzz.token_sort_ratio(brand_name.lower(), inner_item['name'].lower()) >= 50 \
                    and inner_item['name'] not in similar_brands:
                similar_brands.append(inner_item['name'])
            # Поиск похожих кириллических брендов
            elif 'cyrillic-name' in inner_item and fuzz.token_sort_ratio(brand_name.lower(),
                                                                         inner_item['cyrillic-name'].lower()) >= 50 \
                    and inner_item['name'] not in similar_brands:
                similar_brands.append(inner_item['name'])
        if similar_brands:
            response_message = "Похожие бренды:\n" + "\n".join(similar_brands)
            await event.answer(response_message)
    return models


# Команды
@router.message(F.text == "Перезагрузить бота")
@router.message(F.text == "Добавить ещё объявление")
@router.message(F.text == "Отменить и заполнить заново")
@router.message(User.STATE_SUPPORT_END)
@router.message(Command("restart"))
async def restart(message: types.Message, state: FSMContext):
    await state.clear()
    await message.answer("Бот перезапущен.")
    await start(message, state)


@router.message(Command("support"))
async def support(message: types.Message, state: FSMContext):
    await state.clear()
    secret_number = str(random.randint(100, 999))

    await message.answer(f"Нашли баг? Давайте отправим сообщение разработчикам! "
                         f"Но перед этим введите проверку. Докажите что вы не робот. Напишите число {secret_number}:")
    await state.set_state(User.STATE_SUPPORT_VALIDATION)


@router.message(User.STATE_SUPPORT_VALIDATION)
async def support_validation(message: types.Message, state: FSMContext, secret_number):
    if message.text.isdigit() and message.text == secret_number:
        await message.reply(f"Проверка пройдена успешно!")
        await asyncio.sleep(1)
        await message.answer(f"Опишите техническую проблему в деталях для разработчиков: ")
        await state.set_state(User.STATE_SUPPORT_MESSAGE)
    else:
        await message.answer(f"Попробуйте ещё раз!")
        await asyncio.sleep(1)
        await support(message, state)


@router.message(User.STATE_SUPPORT_MESSAGE)
async def support_message(message: types.Message, state: FSMContext):
    current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    message_to_write = f"""
    Дата: {current_time}
    Имя: {message.from_user.full_name}
    Telegram @{message.from_user.username or message.from_user.id} 
  
    Сообщение: {message.text}
    ...
        """

    # Открываем файл для записи и записываем сообщение
    with open("support.txt", "a") as file:
        file.write(message_to_write)
    builder = create_keyboard(['Перезагрузить бота'])
    await message.reply("Спасибо за ваше сообщение! Мы рассмотрим вашу проблему!",
                        reply_markup=builder.as_markup(resize_keyboard=True))
    await state.set_state(User.STATE_SUPPORT_END)


# @router.message(User.STATE_SUPPORT_END)
# async def support_end(message: types.Message, state: FSMContext):
#     await restart(message, state)


# Начало работы бота

@router.message(CommandStart())
async def start(message: types.Message, state: FSMContext):
    image_hello_path = ImageDirectory.auto_say_hi
    await message.answer_photo(photo=types.FSInputFile(image_hello_path),
                               caption=f"Привет, {message.from_user.first_name}! Давай продадим твоё авто! Начнём же сбор данных!")
    await asyncio.sleep(0.1)
    keyboard = create_keyboard(list(dict_car_brands_and_models.keys()))
    image_path = ImageDirectory.auto_car_brand
    await message.answer_photo(photo=types.FSInputFile(image_path), caption="Выберите бренд автомобиля:",
                               reply_markup=keyboard.as_markup())
    await state.set_state(User.STATE_CAR_BRAND)


@router.message(User.STATE_CAR_BRAND)
async def get_car_brand(message: types.Message, state: FSMContext):
    user_data = (await state.get_data()).get("user_data", {})
    await state.set_state(User.STATE_CAR_PHOTO)


@router.message(User.STATE_CAR_PHOTO)
@router.message(F.media_group_id)
async def handle_photos(event: types.Message, state: FSMContext, album: list[Message]):
    user_data = await state.get_data()
    if 'sent_photos' not in user_data:
        user_data['sent_photos'] = []

    new_id = str(uuid.uuid4().int)[:6]

    caption = (
        f"💬<b>Телеграм:</b> <span class='tg-spoiler'>{event.from_user.username if event.from_user.username is not None else 'по номеру телефона'}</span>\n\n"
        f"<b>ID объявления: #{new_id}</b>"
    )

    for message in album:
        # Проверяем, есть ли атрибут photo у текущего сообщения
        if message.photo:
            # Берем последний элемент списка photo, который обычно является наивысшим качеством
            top_photo = message.photo[-1]
            # Получаем file_id фотографии и добавляем его в список sent_photos
            user_data['sent_photos'].append(
                InputMediaPhoto(media=top_photo.file_id, caption=None, parse_mode="HTML"))

    user_data['sent_photos'][0].caption = caption

    await state.update_data(user_data)

    builder = ReplyKeyboardBuilder([[types.KeyboardButton(text="Следущий шаг"), ]])
    if album:
        count_photos = len(album)
        await event.reply(f'{count_photos} Фото добавлены', reply_markup=builder.as_markup(resize_keyboard=True))

    await state.set_state(User.STATE_PREVIEW_ADVERTISMENT)


@router.message(F.text == "Отправить в канал")
async def send_advertisement(message: types.Message, state):
    user_data = await state.get_data()
    print('2', user_data['sent_photos'])

    await add_data_to_excel(message, state)
    user_id = message.from_user.id
    await bot.send_media_group(chat_id=CHANNEL_ID, media=user_data['sent_photos'], disable_notification=True)
    builder = create_keyboard(['Добавить ещё объявление', 'Ускорить продажу'])
    await bot.send_message(user_id, "Объявление отправлено в канал!",
                           reply_markup=builder.as_markup(resize_keyboard=True))
    await state.clear()


# @router.message(F.text == "Отменить и заполнить заново")
# async def fill_again(message: types.Message, state: FSMContext):
#     user_data = await state.get_data()
#     builder = create_keyboard(list(dict_car_brands_and_models.keys()))
#     image_path = ImageDirectory.auto_car_brand
#     with open(image_path, "rb"):
#         await message.answer_photo(photo=types.FSInputFile(image_path), caption="Выберите бренд автомобиля:",
#                                    reply_markup=builder.as_markup(resize_keyboard=True, row_width=2))
#     user_data['sent_photos'].clear()
#     await state.clear()
#     await state.set_state(User.STATE_CAR_BRAND)


@router.message(F.text == "Ускорить продажу")
async def promotion(message: types.Message):
    builder = create_keyboard(['Перезагрузить бота'])
    await message.reply("Чтобы купить закреп напишите @selbie_adv",
                        reply_markup=builder.as_markup(resize_keyboard=True))


@router.message(User.STATE_PREVIEW_ADVERTISMENT)
async def preview_advertisement(message: types.Message, state: FSMContext):
    user_data = await state.get_data()
    print('1', user_data['sent_photos'])
    await bot.send_media_group(chat_id=message.chat.id, media=user_data['sent_photos'])

    builder = ReplyKeyboardBuilder([[
        KeyboardButton(text="Отправить в канал"),
        KeyboardButton(text="Отменить и заполнить заново")
    ]])

    await message.reply(
        "Так будет выглядеть ваше объявление. Вы можете либо разместить либо отменить и заполнить заново.",
        reply_markup=builder.as_markup(resize_keyboard=True))

    # db_fix.clear()


async def add_data_to_excel(message, state):
    user_data = await state.get_data()
    print('3', user_data['sent_photos'])
    file_path = 'db.xlsx'
    row_data = [
        # db_fix.get('new_id'),
        datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        # user_data.get('car_brand', ''),
        message.from_user.username if message.from_user.username is not None else 'по номеру телефона',
    ]

    # Проверяем, существует ли файл Excel
    if os.path.exists(file_path):
        workbook = openpyxl.load_workbook(file_path)
    else:
        workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Проверяем, нужно ли добавить заголовки
    if sheet.max_row == 1:
        headers = [
            'ID', 'Дата', 'Бренд', 'Модель', 'Год', 'Пробег (км)', 'Тип трансмиссии',
            'Тип кузова', 'Тип двигателя', 'Объем двигателя (л)', 'Мощность (л.с.)',
            'Цвет', 'Статус документа', 'Количество владельцев', 'Растаможен',
            'Состояние', 'Дополнительное описание', 'Цена', 'Валюта',
            'Местоположение', 'Имя продавца', 'Телефон продавца', 'Телеграм'
        ]
        sheet.append(headers)

    sheet.append(row_data)
    workbook.save(file_path)


# end support


# старт бота
if __name__ == '__main__':
    asyncio.run(main())
